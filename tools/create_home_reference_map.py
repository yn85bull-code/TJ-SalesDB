from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.comments import Comment

OUT = (
    r"C:\Users\user\Documents\Codex\2026-05-31"
    r"\google-drive-plugin-google-drive-openai\outputs"
    r"\TJ-SalesOS_HOME参照マッピング記入表.xlsx"
)

ROWS = [
    [1, "最上部KPI", "TOPLINE", "TOPLINE", "V2", "セル確定", "", "11桁想定。OS表示は全桁表示。"],
    [2, "最上部KPI", "G", "G", "E2", "セル確定", "", ""],
    [3, "最上部KPI", "MQ", "MQ", "D2", "セル確定", "", ""],
    [4, "最上部KPI", "F", "F", "C2", "セル確定", "", ""],
    [5, "買取実績", "買取台数", "買取Q / 買取台数実績", "C6", "セル確定", "", "表示単位は台。"],
    [6, "買取実績", "買取Q前日差", "買取Q前日差 / 買取前日差", "", "要確認", "", "前日比グラフ・増減表示に使用。"],
    [7, "買取実績", "査定Q", "査定Q / 査定台数実績", "B6", "セル確定", "", "表示単位は台。"],
    [8, "買取実績", "成約CVR", "成約CVR / 買取成約CVR", "E6", "セル確定", "", "比率グラフ表示。"],
    [9, "買取実績", "AAQ", "AAQ", "B8", "セル確定", "", "販路内訳。母数は買取Q。"],
    [10, "買取実績", "商品Q", "商品Q", "C8", "セル確定", "", "販路内訳。母数は買取Q。"],
    [11, "買取実績", "スクラップQ", "スクラップQ", "D8", "セル確定", "", "販路内訳。母数は買取Q。"],
    [12, "買取実績", "代車Q", "代車Q", "E8", "セル確定", "", "販路内訳。母数は買取Q。"],
    [13, "買取実績", "出品Q", "出品Q / 当月AA出品Q / 当月買取出品Q", "", "要確認", "", "出品Q母数で落札比率を表示。"],
    [14, "買取実績", "落札Q", "落札Q / AA落札Q", "B14", "セル確定", "", ""],
    [15, "買取実績", "流札Q", "流札Q / AA流札Q", "C14", "セル確定", "", ""],
    [16, "買取実績", "取消Q", "取消Q / AA取消Q", "D14", "セル確定", "", ""],
    [17, "買取実績", "落札MQ", "落札MQ / AAMQ", "B16", "セル確定", "", "AA収益。"],
    [18, "買取実績", "AA@", "AA@ / AA＠", "C16", "セル確定", "", ""],
    [19, "買取実績", "翌月Q", "持越Q / AA翌月Q", "F14", "セル確定", "", ""],
    [20, "販売実績", "販売Q", "販売Q / 販売台数実績", "", "要確認", "", "表示単位は台。"],
    [21, "販売実績", "販売Q前日差", "販売Q前日差 / 販売前日差", "", "要確認", "", "前日比グラフ・増減表示に使用。"],
    [22, "販売実績", "在庫Q", "在庫Q / 在庫台数", "B20", "セル確定", "", ""],
    [23, "販売実績", "未掲載Q", "未掲載Q / 未掲載数", "F20", "セル確定", "", "母数は在庫Q。GAS/HTML側で比率計算。"],
    [24, "販売実績", "未受入Q", "未受入Q / 未受入数", "H20", "セル確定", "", "母数は在庫Q。GAS/HTML側で比率計算。"],
    [25, "販売実績", "未入庫Q", "未入庫Q / 未入庫台数", "G20", "セル確定", "", "母数は配車在庫Q。GAS/HTML側で比率計算。"],
    [26, "販売実績", "配車在庫Q", "配車在庫Q / 配車在庫数", "I20", "セル確定", "", "在庫状況ドーナツの母数。"],
    [27, "販売実績", "納車Q", "納車Q / 当月納車Q", "B22", "セル確定", "", ""],
    [28, "販売実績", "未納車Q", "未納車Q", "", "要確認", "", "納車Qとの比率棒グラフに使用。"],
    [29, "販売実績", "回転率予想", "回転率予想 / 回転数予想", "Q25想定", "要確認", "", "目標対比グラフ。"],
    [30, "販売実績", "回転率目標", "回転率目標 / 回転数目標", "P25想定", "要確認", "", "目標対比グラフ。"],
    [31, "販売実績", "販売MQ", "販売総MQ", "J24", "セル確定", "", "販売収益。"],
    [32, "販売実績", "販売@", "販売総MQ ÷ 販売Q", "HTML計算", "計算", "", "販売MQを販売Qで割る。"],
    [33, "販売実績", "翌月納車Q", "翌月納車Q", "D22想定", "要確認", "", ""],
    [34, "集客情報", "情報数", "情報数", "M6", "セル確定", "", ""],
    [35, "集客情報", "情報数前日差", "情報数前日差", "", "要確認", "", "前日比グラフ・増減表示に使用。"],
    [36, "集客情報", "情報料金", "情報料金", "", "要確認", "", ""],
    [37, "集客情報", "開示率", "開示率", "", "要確認", "", "3本グラフ表示。"],
    [38, "集客情報", "査定率", "査定率", "", "要確認", "", "3本グラフ表示。"],
    [39, "集客情報", "成約率", "成約率", "", "要確認", "", "3本グラフ表示。"],
    [40, "集客情報", "情報数目標", "情報数目標", "P26", "セル確定", "", ""],
    [41, "集客情報", "情報数予想", "情報数予想", "Q26", "セル確定", "", ""],
    [42, "集客情報", "開示数目標", "開示数目標", "P27", "セル確定", "", ""],
    [43, "集客情報", "開示数予想", "開示数予想", "Q27", "セル確定", "", ""],
    [44, "集客情報", "査定数目標", "査定数目標", "P28", "セル確定", "", ""],
    [45, "集客情報", "査定数予想", "査定数予想", "Q28", "セル確定", "", ""],
    [46, "集客情報", "成約数目標", "成約数目標", "P29", "セル確定", "", ""],
    [47, "集客情報", "成約数予想", "成約数予想", "Q29", "セル確定", "", ""],
    [48, "集客情報", "情報成約率", "成約数 ÷ 情報数", "HTML計算", "計算", "", ""],
    [49, "集客情報", "開示情報成約率", "成約数 ÷ 開示数", "HTML計算", "計算", "", ""],
    [50, "目標進捗", "G目標", "G目標", "P15", "セル確定", "", ""],
    [51, "目標進捗", "MQ目標", "MQ目標", "P16", "セル確定", "", ""],
    [52, "目標進捗", "買取Q目標", "買取Q目標", "P17", "セル確定", "", ""],
    [53, "目標進捗", "販売Q目標", "販売Q目標", "P18", "セル確定", "", ""],
    [54, "目標進捗", "落札MQ目標", "落札MQ目標", "P19", "セル確定", "", ""],
    [55, "目標進捗", "販売総MQ目標", "販売総MQ目標", "P20", "セル確定", "", ""],
    [56, "目標進捗", "査定CVR目標", "査定CVR目標", "P21", "セル確定", "", ""],
    [57, "目標進捗", "成約CVR目標", "成約CVR目標", "P22", "セル確定", "", ""],
    [58, "目標進捗", "AA@目標", "AA@目標", "P23", "セル確定", "", ""],
    [59, "目標進捗", "納車@目標", "納車@目標", "P24", "セル確定", "", ""],
    [60, "個人ランキング", "MQランキング", "個人実績表", "A31:L63", "表範囲", "", "1-3位表示。"],
    [61, "個人ランキング", "買取ランキング", "個人実績表", "A31:L63", "表範囲", "", "1-3位表示。"],
    [62, "個人ランキング", "販売ランキング", "個人実績表", "A31:L63", "表範囲", "", "1-3位表示。"],
    [63, "カレンダー", "日別 買取Q / 販売Q / 成約CVR", "カレンダー表", "N31:R62", "表範囲", "", "月末まで自動表示。31日月も対応想定。"],
]


def style_sheet(ws, headers, widths):
    red = "D9001B"
    black = "111111"
    white = "FFFFFF"
    line = Side(style="thin", color="D9DDE5")
    medium_red = Side(style="medium", color=red)
    header_fill = PatternFill("solid", fgColor=black)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color=white, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=medium_red, bottom=medium_red, left=line, right=line)

    section_colors = {
        "最上部KPI": "F4CCCC",
        "買取実績": "E2F0D9",
        "販売実績": "D9EAF7",
        "集客情報": "EADCF8",
        "目標進捗": "FCE4D6",
        "個人ランキング": "F4CCCC",
        "カレンダー": "E7E6E6",
    }
    status_colors = {
        "セル確定": "E2F0D9",
        "要確認": "FFF2CC",
        "計算": "DDEBF7",
        "表範囲": "E7E6E6",
    }
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        section = row[1].value
        status = row[5].value
        section_fill = PatternFill("solid", fgColor=section_colors.get(section, white))
        for i, cell in enumerate(row, start=1):
            cell.border = Border(top=line, bottom=line, left=line, right=line)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if i in (1, 2):
                cell.fill = section_fill
                cell.font = Font(bold=True)
            if i == 6:
                cell.fill = PatternFill("solid", fgColor=status_colors.get(status, white))
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if i == 7:
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
                cell.comment = Comment(
                    "ここに正しいセル番地・範囲・メモを記入してください。確定済みでも修正があれば上書きOKです。",
                    "Codex",
                )

    ws.conditional_formatting.add(
        f"A2:H{ws.max_row}",
        FormulaRule(formula=['$F2="要確認"'], fill=PatternFill("solid", fgColor="FFF2CC")),
    )
    status_dv = DataValidation(type="list", formula1='"セル確定,要確認,計算,表範囲"', allow_blank=False)
    ws.add_data_validation(status_dv)
    status_dv.add(f"F2:F{ws.max_row}")
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{ws.max_row}"


def main():
    wb = Workbook()
    headers = ["No", "表示場所", "表示項目", "参照キー", "現在の参照セル/範囲", "状態", "ユーザー記入セル", "備考"]
    widths = [6, 16, 24, 36, 22, 12, 24, 46]

    ws = wb.active
    ws.title = "HOME参照マップ"
    ws.append(headers)
    for row in ROWS:
        ws.append(row)
    style_sheet(ws, headers, widths)

    rule_ws = wb.create_sheet("記入ルール")
    rule_ws.append(["TJ-SalesOS HOME参照マッピング記入ルール"])
    rule_ws.append([""])
    for row in [
        ["1", "黄色の『ユーザー記入セル』に、正しいセル番地・範囲・補足を記入してください。"],
        ["2", "『状態』が要確認の行を優先して埋めてください。"],
        ["3", "既にセル確定になっている行でも、間違いがあれば『ユーザー記入セル』に正しい参照先を書いてOKです。"],
        ["4", "HTML計算の行は、元になる項目が正しければ基本そのままでOKです。"],
        ["5", "表範囲の行は、ランキングやカレンダーなど複数行の範囲指定です。"],
        ["6", "空白の値は0表示、参照項目自体がない場合だけ未取得扱いにする想定です。"],
    ]:
        rule_ws.append(row)
    rule_ws["A1"].font = Font(size=16, bold=True, color="D9001B")
    rule_ws["A1"].fill = PatternFill("solid", fgColor="FDE8EC")
    rule_ws.merge_cells("A1:B1")
    line = Side(style="thin", color="D9DDE5")
    for row in rule_ws.iter_rows(min_row=3, max_row=8):
        for cell in row:
            cell.border = Border(top=line, bottom=line, left=line, right=line)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    rule_ws.column_dimensions["A"].width = 8
    rule_ws.column_dimensions["B"].width = 90

    check_ws = wb.create_sheet("要確認だけ")
    check_ws.append(headers)
    for row in ROWS:
        if row[5] == "要確認":
            check_ws.append(row)
    style_sheet(check_ws, headers, widths)

    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
