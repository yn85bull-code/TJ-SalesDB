import json
from pathlib import Path
from openpyxl import load_workbook

SRC = r"C:\Users\user\Downloads\営業部ダッシュボード (1).xlsx"
SHEET_NAME = "OS買取詳細用"
OUT_HTML = Path("outputs/tj-sales-os-lite-dashboard.html")
OUT_GAS = Path("outputs/tj-sales-os-lite-Code.gs")

DISPLAY_KEYS = [
    "営業部名",
    "査定Ｑ", "買取Ｑ", "成約CVR",
    "AA@", "AAQ", "商品Q", "スクラップQ", "代車Q",
    "AA比率", "商品比率", "スクラップ比率", "代車比率",
    "買取キャンセルQ", "キャンセル比率", "入庫平均日数", "即決契約数", "管理契約数",
    "当月AA出品Q", "当月出品Q", "当月出品比率",
    "AA相違Q", "AA相違比率", "当月AAMQ",
    "AA落札Q", "AA未処理Q", "AA流札Q", "AA取消Q", "AA翌月Q",
    "AA赤字金額", "AA赤字Q", "AA赤字比率",
]

wb = load_workbook(SRC, data_only=True)
ws = wb[SHEET_NAME]
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
rows = []
for r in range(2, ws.max_row + 1):
    obj = {}
    empty = True
    for c, h in enumerate(headers, start=1):
        if not h:
            continue
        v = ws.cell(r, c).value
        if v not in (None, ""):
            empty = False
        obj[str(h)] = v
    if not empty and obj.get("営業部名"):
        rows.append({k: obj.get(k, "") for k in DISPLAY_KEYS})

sample_json = json.dumps(rows, ensure_ascii=False)

html = f"""<!doctype html>
<html lang="ja">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>TJ-SalesOS 買取詳細</title>
  <style>
    :root {{
      --bg:#f4f4f4;
      --line:#050505;
      --red:#d4001a;
      --red-dark:#8a0012;
      --blue:#1085c7;
      --yellow:#f2b51d;
      --ink:#050505;
      --muted:#555;
      --pink:linear-gradient(100deg,#fff 0%,#fff 45%,#ffe9ed 73%,#ffc9d3 100%);
    }}
    * {{ box-sizing:border-box; }}
    body {{
      margin:0;
      background:var(--bg);
      color:var(--ink);
      font-family:"Yu Gothic","Meiryo","Segoe UI",sans-serif;
      font-size:12px;
    }}
    .shell {{ min-height:100vh; display:grid; grid-template-columns:220px minmax(1120px,1fr); align-items:start; }}
    .side {{
      background:#060606; color:#fff; padding:18px 14px; position:sticky; top:0; height:100vh;
      border-right:3px solid #111;
    }}
    .brand {{ font-size:22px; font-weight:1000; line-height:1; letter-spacing:.04em; margin-bottom:5px; }}
    .brand-sub {{ color:#d7d7d7; font-size:10px; font-weight:900; letter-spacing:.08em; margin-bottom:18px; }}
    .side .dept-select {{ width:100%; min-width:0; background:#111; color:#fff; border:2px solid #fff; margin-bottom:16px; }}
    .menu {{ display:grid; gap:8px; }}
    .menu button {{
      border:2px solid #2c2c2c; background:#111; color:#e5e5e5; padding:11px 10px; text-align:left;
      font-weight:1000; cursor:pointer; border-left:5px solid transparent;
    }}
    .menu button.active {{ color:#fff; border-color:#fff; border-left-color:var(--red); background:#1a1a1a; }}
    .page {{ width:100%; max-width:none; margin:0; padding:14px 22px 34px; }}
    .crumb {{ margin:0 0 10px; color:#444; font-weight:800; }}
    .title-bar {{
      display:flex; align-items:center; justify-content:space-between; gap:16px;
      height:45px; border:3px solid var(--line); border-left:7px solid var(--red);
      background:#fff; padding:0 14px; margin-bottom:12px;
    }}
    .title-bar h1 {{ margin:0; font-size:18px; font-weight:1000; }}
    .dept-select {{ min-width:190px; border:2px solid var(--line); background:#fff; padding:6px 10px; font-weight:1000; }}
    .panel {{ border:3px solid var(--line); background:#fff; padding:11px; margin-bottom:14px; }}
    .panel-title {{
      margin:0 0 10px; padding-bottom:8px; border-bottom:3px solid var(--line);
      font-size:13px; font-weight:1000; color:var(--red);
    }}
    .grid {{ display:grid; gap:8px; }}
    .top-kpi {{ grid-template-columns:1fr 1fr; }}
    .split-2 {{ grid-template-columns:1fr 1fr; }}
    .split-3 {{ grid-template-columns:1.05fr .95fr; }}
    .mini-4 {{ grid-template-columns:repeat(4,minmax(0,1fr)); }}
    .aa-grid {{ grid-template-columns:.82fr 1.18fr; align-items:stretch; }}
    .box {{
      border:2px solid var(--line); background:#fff; padding:10px; min-height:68px;
      position:relative; overflow:hidden;
    }}
    .box.tint, .big-kpi, .aa-main {{ background:var(--pink); }}
    .box-label {{ font-size:11px; font-weight:1000; color:#222; margin-bottom:6px; }}
    .flag {{
      display:inline-flex; align-items:center; justify-content:center;
      min-width:47px; height:23px; background:linear-gradient(90deg,#930012,#d4001a);
      color:#fff; border:1px solid #73000e; font-size:11px; font-weight:1000; padding:0 8px;
      box-shadow:1px 1px 0 #111;
    }}
    .big-kpi {{ min-height:112px; display:flex; flex-direction:column; justify-content:space-between; }}
    .big-row {{ display:flex; align-items:flex-end; justify-content:space-between; gap:12px; }}
    .big-value {{ font-size:48px; line-height:.95; font-weight:1000; letter-spacing:-.02em; white-space:nowrap; }}
    .unit {{ font-size:.62em; margin-left:2px; }}
    .black-badge {{
      min-width:116px; border:2px solid var(--red); background:#050505; color:#fff; padding:10px;
      text-align:right; box-shadow:3px 3px 0 rgba(0,0,0,.25);
    }}
    .black-badge span {{ display:block; font-size:10px; font-weight:1000; margin-bottom:3px; }}
    .black-badge b {{ font-size:29px; line-height:1; }}
    .metric-value {{ font-size:24px; line-height:1; font-weight:1000; white-space:nowrap; }}
    .metric-value.negative {{ color:var(--red); }}
    .bar-line {{ margin-top:8px; }}
    .bar-line .bar-meta {{ display:flex; justify-content:space-between; gap:10px; font-weight:900; font-size:10px; margin-bottom:3px; }}
    .track {{ height:8px; border:1px solid #111; background:#fff; border-radius:9px; overflow:hidden; }}
    .fill {{ height:100%; width:0; background:linear-gradient(90deg,var(--red-dark),#111); border-right:1px solid #111; transition:width .85s ease; }}
    .stack {{ border:2px solid var(--line); background:#fff; padding:10px; }}
    .stack-title {{ margin:0 0 8px; color:var(--red); font-weight:1000; }}
    .stackbar {{ height:15px; border:2px solid #111; background:#fff; border-radius:10px; overflow:hidden; display:flex; }}
    .seg {{ height:100%; min-width:0; }}
    .seg.red {{ background:var(--red); }}
    .seg.black {{ background:#111; }}
    .seg.blue {{ background:var(--blue); }}
    .seg.yellow {{ background:var(--yellow); }}
    .legend {{ display:grid; grid-template-columns:1fr 1fr; gap:5px 18px; margin-top:8px; font-size:10px; font-weight:800; }}
    .dot {{ display:inline-block; width:7px; height:7px; border-radius:50%; border:1px solid #111; margin-right:5px; vertical-align:middle; }}
    .dot.red {{ background:var(--red); }}
    .dot.black {{ background:#111; }}
    .dot.blue {{ background:var(--blue); }}
    .dot.yellow {{ background:var(--yellow); }}
    .aa-main {{ min-height:244px; display:grid; grid-template-rows:1fr auto auto; gap:8px; }}
    .aa-main .big-value {{ font-size:66px; }}
    .aa-subgrid {{ display:grid; grid-template-columns:1fr 1fr; gap:7px; }}
    .aa-side {{ display:grid; grid-template-columns:1fr; gap:8px; }}
    .aa-counts {{ display:grid; grid-template-columns:repeat(3,1fr); gap:7px; }}
    .warn-grid {{ display:grid; grid-template-columns:repeat(3,1fr); gap:7px; padding-top:8px; border-top:3px solid #111; }}
    .warn {{ border:2px solid #d6a40d; background:#fff9df; padding:9px; }}
    .warn .metric-value {{ font-size:22px; }}
    @media (max-width:720px) {{
      .shell {{ grid-template-columns:170px minmax(0,1fr); }}
      .side {{ padding:14px 10px; }}
      .brand {{ font-size:18px; }}
      .menu button {{ padding:9px 8px; font-size:11px; }}
      .page {{ padding:12px 14px 30px; }}
      .title-bar {{ height:auto; min-height:44px; padding:9px 12px; }}
      .top-kpi,.split-2,.split-3,.aa-grid,.aa-counts,.warn-grid {{ grid-template-columns:1fr; }}
      .mini-4 {{ grid-template-columns:1fr 1fr; }}
      .big-kpi {{ min-height:116px; }}
      .big-value {{ font-size:42px; }}
      .black-badge {{ min-width:96px; padding:8px; }}
      .black-badge b {{ font-size:22px; }}
      .metric-value {{ font-size:21px; }}
      .aa-main .big-value {{ font-size:52px; }}
    }}
  </style>
</head>
<body>
  <div class="shell">
    <aside class="side">
      <div class="brand">TJ-SalesOS</div>
      <div class="brand-sub">PURCHASE DETAIL</div>
      <select id="deptSelect" class="dept-select"></select>
      <nav class="menu">
        <button type="button">⌂ HOME</button>
        <button type="button">♡ 実績管理</button>
        <button type="button" class="active">♤ 買取詳細</button>
        <button type="button">♢ 販売詳細</button>
        <button type="button">♧ 社員管理</button>
      </nav>
    </aside>
    <div class="page">
      <p class="crumb">買取詳細 / <span id="crumbDept">-</span> / 更新完了</p>
      <div class="title-bar">
        <h1>買取詳細</h1>
        <div style="font-weight:1000;" id="titleDept">-</div>
      </div>
      <div id="content"></div>
    </div>
  </div>
  <script>
    const MOCK_ROWS = {sample_json};
    let rows = [];
    let current = null;

    const moneyKKeys = new Set(['AA@','当月AAMQ','AA赤字金額']);
    const percentKeys = new Set(['成約CVR','AA比率','商品比率','スクラップ比率','代車比率','当月出品比率','AA相違比率','キャンセル比率','AA赤字比率']);

    function runGas(name, args) {{
      return new Promise((resolve, reject) => {{
        if (!window.google || !google.script || !google.script.run) return resolve(MOCK_ROWS);
        google.script.run.withSuccessHandler(resolve).withFailureHandler(reject)[name].apply(google.script.run, args || []);
      }});
    }}
    function esc(v) {{ return String(v ?? '').replace(/[&<>"']/g, s => ({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#039;'}}[s])); }}
    function num(v) {{ const n = Number(String(v ?? '').replace(/[¥￥,%台件日\\s]/g,'')); return isNaN(n) ? 0 : n; }}
    function pct(v) {{ const n = num(v); return Math.abs(n) <= 1 ? n * 100 : n; }}
    function fmtPct(v) {{ return `${{pct(v).toFixed(1)}}%`; }}
    function fmtValue(key, value) {{
      if (value === '' || value === null || value === undefined) return '-';
      const n = num(value);
      if (percentKeys.has(key)) return fmtPct(value);
      if (moneyKKeys.has(key)) return `${{n < 0 ? '-' : ''}}¥${{Math.abs(Math.round(n * 1000)).toLocaleString('ja-JP')}}`;
      if (key === '入庫平均日数') return `${{n.toFixed(1).replace(/\\.0$/,'')}}`;
      return Math.round(n).toLocaleString('ja-JP');
    }}
    function signedClass(v) {{ return num(v) < 0 ? 'negative' : ''; }}
    function metric(label, key, suffix='', tint=false) {{
      const value = current?.[key];
      return `<div class="box ${{tint ? 'tint' : ''}}">
        <div class="box-label">${{esc(label)}}</div>
        <div class="metric-value ${{signedClass(value)}}">${{esc(fmtValue(key,value))}}${{suffix}}</div>
      </div>`;
    }}
    function lineMetric(label, key, totalKey, caption) {{
      const v = current?.[key];
      const rate = totalKey ? ratio(v, current?.[totalKey]) : pct(current?.[key]);
      return `<div class="box">
        <div class="box-label">${{esc(label)}}</div>
        <div class="metric-value">${{esc(fmtValue(key,v))}} / ${{rate.toFixed(1)}}%</div>
        <div class="bar-line">
          <div class="bar-meta"><span>${{esc(caption || label)}}</span><span>${{rate.toFixed(1)}}%</span></div>
          <div class="track"><div class="fill" style="width:${{clamp(rate)}}%"></div></div>
        </div>
      </div>`;
    }}
    function ratio(v, total) {{
      const t = num(total);
      if (!t) return 0;
      return num(v) / t * 100;
    }}
    function clamp(n) {{ return Math.max(0, Math.min(100, n || 0)); }}
    function stack(title, items, totalKey) {{
      const total = totalKey ? num(current?.[totalKey]) : items.reduce((s, it) => s + num(current?.[it.key]), 0);
      const colors = ['red','black','blue','yellow'];
      const segs = items.map((it, i) => `<span class="seg ${{colors[i % colors.length]}}" style="width:${{clamp(total ? num(current?.[it.key]) / total * 100 : 0)}}%"></span>`).join('');
      const legend = items.map((it, i) => `<span><i class="dot ${{colors[i % colors.length]}}"></i>${{esc(it.label)}} ${{esc(fmtValue(it.key,current?.[it.key]))}} / ${{(total ? num(current?.[it.key]) / total * 100 : 0).toFixed(1)}}%</span>`).join('');
      return `<div class="stack">
        <p class="stack-title">${{esc(title)}}</p>
        <div class="stackbar">${{segs}}</div>
        <div class="legend">${{legend}}</div>
      </div>`;
    }}
    function bigKpi(label, key, unit, badgeLabel, badgeKey) {{
      return `<div class="big-kpi box">
        <span class="flag">${{esc(label)}}</span>
        <div class="big-row">
          <div class="big-value">${{esc(fmtValue(key,current?.[key]))}}<span class="unit">${{esc(unit)}}</span></div>
          <div class="black-badge"><span>${{esc(badgeLabel)}}</span><b>${{esc(fmtValue(badgeKey,current?.[badgeKey]))}}</b></div>
        </div>
      </div>`;
    }}
    function renderPurchase() {{
      document.getElementById('content').innerHTML = `
        <section class="panel">
          <h2 class="panel-title">買取実績詳細</h2>
          <div class="grid top-kpi">
            ${{bigKpi('査定数','査定Ｑ','件','査定CVR','成約CVR')}}
            ${{bigKpi('買取台数','買取Ｑ','台','成約CVR','成約CVR')}}
          </div>
          <div class="grid split-3" style="margin-top:8px;">
            <div class="grid mini-4">
              ${{lineMetric('AA','AAQ','買取Ｑ','AA比率')}}
              ${{lineMetric('商品','商品Q','買取Ｑ','商品比率')}}
              ${{lineMetric('スクラップ','スクラップQ','買取Ｑ','スクラップ比率')}}
              ${{lineMetric('代車','代車Q','買取Ｑ','代車比率')}}
            </div>
            <div class="grid">
              ${{stack('販路構成',[
                {{label:'AA', key:'AAQ'}},
                {{label:'商品', key:'商品Q'}},
                {{label:'スクラップ', key:'スクラップQ'}},
                {{label:'代車', key:'代車Q'}}
              ], '買取Ｑ')}}
            </div>
          </div>
          <div class="grid split-2" style="margin-top:8px;">
            ${{lineMetric('買取キャンセル','買取キャンセルQ','買取Ｑ','買取キャンセル比率')}}
            ${{metric('入庫平均日数','入庫平均日数','日',true)}}
          </div>
        </section>

        <section class="panel">
          <h2 class="panel-title" style="color:#111;">AA出品詳細</h2>
          <div class="grid aa-grid">
            <div class="aa-main box">
              <div>
                <div class="box-label">出品Q</div>
                <div class="big-value">${{esc(fmtValue('当月AA出品Q', current?.['当月AA出品Q']))}}<span class="unit">台</span></div>
              </div>
              <div class="aa-subgrid">
                ${{metric('AAMQ','当月AAMQ')}}
                ${{metric('AA@','AA@')}}
              </div>
              <div class="aa-subgrid">
                ${{metric('処理予定台数','AA未処理Q','台')}}
                ${{metric('処理予定MQ','当月AAMQ')}}
              </div>
            </div>
            <div class="aa-side">
              <div class="aa-counts">
                ${{metric('落札Q','AA落札Q','台')}}
                ${{metric('流札Q','AA流札Q','台')}}
                ${{metric('取消Q','AA取消Q','台')}}
              </div>
              ${{stack('出品Q内訳',[
                {{label:'落札', key:'AA落札Q'}},
                {{label:'流札', key:'AA流札Q'}},
                {{label:'取消', key:'AA取消Q'}}
              ], '当月AA出品Q')}}
              <div class="warn-grid">
                ${{metric('赤字Q比率','AA赤字比率','')}}
                ${{metric('赤字金額','AA赤字金額','')}}
                ${{metric('赤字なしAAMQ','当月AAMQ','')}}
              </div>
            </div>
          </div>
          <div class="grid split-2" style="margin-top:8px;">
            <div class="grid mini-4">
              ${{lineMetric('当月買取出品Q','当月出品Q','買取Ｑ','当月出品比率')}}
              ${{lineMetric('AA相違','AA相違Q','当月AA出品Q','AA相違比率')}}
              ${{lineMetric('落札率','AA落札Q','当月AA出品Q','落札率')}}
            </div>
            ${{metric('持越し台数','AA翌月Q','台',true)}}
          </div>
        </section>`;
    }}
    function render() {{
      document.getElementById('crumbDept').textContent = current?.['営業部名'] || '-';
      document.getElementById('titleDept').textContent = current?.['営業部名'] || '-';
      renderPurchase();
    }}
    async function init() {{
      rows = await runGas('getLiteDashboardRows');
      rows = Array.isArray(rows) ? rows.filter(r => r && r['営業部名']) : MOCK_ROWS;
      const select = document.getElementById('deptSelect');
      select.innerHTML = rows.map(r => `<option>${{esc(r['営業部名'])}}</option>`).join('');
      current = rows.find(r => r['営業部名'] === '千葉営業部') || rows[0];
      select.value = current?.['営業部名'] || '';
      select.onchange = () => {{ current = rows.find(r => r['営業部名'] === select.value) || rows[0]; render(); }};
      render();
    }}
    init();
  </script>
</body>
</html>
"""

gas = """const LITE_CONFIG = {
  spreadsheetId: '1uzzA7hHHPS1WaEACwKK82XL8cuEy4LQmAP_jiz7ycoE',
  htmlFileName: 'tj-sales-os-lite-dashboard',
  sheetName: 'OS買取詳細用'
};

function doGet() {
  return HtmlService
    .createHtmlOutputFromFile(LITE_CONFIG.htmlFileName)
    .setTitle('TJ-SalesOS Lite')
    .setXFrameOptionsMode(HtmlService.XFrameOptionsMode.ALLOWALL);
}

function getLiteDashboardRows() {
  const ss = SpreadsheetApp.openById(LITE_CONFIG.spreadsheetId);
  const sheet = ss.getSheetByName(LITE_CONFIG.sheetName);
  if (!sheet) throw new Error(LITE_CONFIG.sheetName + ' が見つかりません。');
  const values = sheet.getDataRange().getDisplayValues();
  if (values.length < 2) return [];
  const headers = values[0].map(v => String(v || '').trim());
  return values.slice(1)
    .filter(row => row.some(cell => cell !== ''))
    .map(row => {
      const obj = {};
      headers.forEach((header, i) => {
        if (header) obj[header] = row[i];
      });
      return obj;
    })
    .filter(row => row['営業部名']);
}
"""

OUT_HTML.write_text(html, encoding="utf-8")
OUT_GAS.write_text(gas, encoding="utf-8")
print(OUT_HTML)
print(OUT_GAS)
