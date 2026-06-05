from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from pathlib import Path

OUTPUT = Path("outputs/TJ-SalesOS_参照マッピング記入表.xlsx")

HEADERS = [
    "No",
    "メニュー",
    "セクション",
    "OS表示項目",
    "単位/型",
    "現在の候補名",
    "記入_参照シート名",
    "記入_列番号",
    "記入_列名",
    "記入_集計条件",
    "記入_例の値",
    "優先度",
    "備考",
]

ROWS = []

def add(menu, section, item, unit="", candidates="", priority="高", note=""):
    ROWS.append([
        len(ROWS) + 1,
        menu,
        section,
        item,
        unit,
        candidates,
        "",
        "",
        "",
        "",
        "",
        priority,
        note,
    ])

# HOME
for item, unit, cand in [
    ("営業部名", "文字", "営業部名 / 営業部 / 部署"),
    ("対象月", "日付/月", "対象月 / 月 / 年月"),
    ("更新日時", "日時", "更新日時 / 最終更新 / 更新日"),
    ("TOPLINE", "金額 千円入力→円表示", "売上 / TOPLINE / 総売上"),
    ("G", "金額 千円入力→円表示", "G / 粗利 / 総粗利"),
    ("F", "金額 千円入力→円表示", "トータルF / F / 総F"),
    ("MQ", "金額 千円入力→円表示", "トータルMQ / MQ / 総MQ"),
    ("査定Q", "件", "査定Q / 査定台数実績 / 査定数"),
    ("買取Q", "台", "買取Q / 買取台数実績 / 買取数"),
    ("販売Q", "台", "販売Q / 販売台数実績 / 納車Q"),
    ("査定CVR", "%", "査定CVR / 査定率"),
    ("成約CVR", "%", "成約CVR / 買取成約CVR / 成約率"),
    ("AA@", "金額 千円入力→円表示", "AA@ / AA平均粗利"),
    ("販売@", "金額 千円入力→円表示", "販売@ / 販売平均粗利"),
    ("AA割合", "%", "AA割合"),
    ("展示割合", "%", "展示割合"),
    ("スタッフ数", "人", "スタッフ数 / スタッフ人数"),
    ("日数進捗", "%", "日数進捗"),
]:
    add("HOME", "トップKPI", item, unit, cand)

# Performance
for item in ["G", "MQ", "査定Q", "買取Q", "販売Q", "査定CVR", "成約CVR"]:
    add("実績管理", "目標進捗", item, "金額/件/台/%", item + " / " + item + "目標", "高", "目標・実績・差分・予想実績・ステータス判定に使用")

for item, unit in [
    ("拠点名", "文字"), ("TOPLINE", "金額"), ("G", "金額"), ("MQ", "金額"), ("F", "金額"),
    ("人数", "人"), ("査定Q", "件"), ("買取Q", "台"), ("販売Q", "台"), ("査定CVR", "%"), ("成約CVR", "%")
]:
    add("実績管理", "営業部内 拠点実績", item, unit, item)

for item, unit in [
    ("氏名", "文字"), ("役職", "文字"), ("ランク", "文字"), ("拠点名", "文字"), ("社歴", "文字"),
    ("査定Q", "件"), ("買取Q", "台"), ("販売Q", "台"), ("査定CVR", "%"), ("成約CVR", "%"), ("MQ", "金額")
]:
    add("実績管理", "営業部内 個人実績/ランキング", item, unit, item)

for item, unit in [
    ("日付", "日付"), ("曜日", "文字"), ("査定Q", "件"), ("買取Q", "台"), ("販売Q", "台"),
    ("査定CVR", "%"), ("成約CVR", "%"), ("情報数", "件")
]:
    add("実績管理", "当月カレンダー", item, unit, item)

for item, unit in [
    ("TOTAL MQ", "金額"), ("買取MQ", "金額"), ("販売MQ", "金額"), ("その他MQ", "金額"),
    ("TOTAL F", "金額"), ("固定費", "金額"), ("人件費", "金額"), ("奨励金", "金額"),
    ("情報料", "金額"), ("掲載料", "金額"), ("販管費", "金額"), ("CC運営費", "金額")
]:
    add("実績管理", "損益実績詳細", item, unit, item)

# Purchase
for item, unit in [
    ("査定数", "件"), ("査定CVR", "%"), ("買取台数", "台"), ("成約CVR", "%"),
    ("AA台数", "台"), ("展示台数", "台"), ("スクラップ台数", "台"), ("代車台数", "台"),
    ("AA比率", "%"), ("展示比率", "%"), ("スクラップ比率", "%"), ("代車比率", "%"),
    ("買取キャンセルQ", "台"), ("キャンセル比率", "%"), ("入庫平均日数", "日"),
    ("即日査定", "件"), ("即日査定成約", "台"), ("過去当日査定", "件"), ("過去当日成約", "台"),
    ("過去後日査定", "件"), ("過去後日成約", "台"), ("後日査定", "件"), ("後日成約", "台"),
]:
    add("買取詳細", "買取実績/査定成約内訳", item, unit, item)

for item, unit in [
    ("出品Q", "台"), ("落札Q", "台"), ("流札Q", "台"), ("取消Q", "台"), ("持越し台数", "台"),
    ("AAMQ", "金額"), ("AA@", "金額"), ("処理予定台数", "台"), ("処理予定MQ", "金額"),
    ("当月買取出品Q", "台"), ("AA相違Q", "台"), ("AA相違比率", "%"),
    ("AA赤字Q", "台"), ("AA赤字比率", "%"), ("AA赤字金額", "金額")
]:
    add("買取詳細", "AA出品詳細", item, unit, item)

for item, unit in [
    ("MOTA情報数", "件"), ("開示数", "件"), ("開示率", "%"), ("査定数", "件"), ("査定率", "%"),
    ("成約数", "台"), ("成約率", "%"), ("情報料金額", "金額"), ("成約CPA", "金額"), ("査定CPA", "金額"),
    ("買取営業人数", "人"), ("Cランク以上", "人"), ("Eランク新人", "人"),
    ("直近層", "件"), ("検討層", "件"), ("潜在層", "件")
]:
    add("買取詳細", "MOTA情報実績", item, unit, item)

# Sales
for item, unit, note in [
    ("在庫台数", "台", "例: 千葉営業部 FC列 7台"),
    ("販売台数", "台", ""),
    ("受注Q", "台", "例: 千葉営業部 AC列 12台"),
    ("納車Q", "台", ""),
    ("未入庫台数", "台", ""),
    ("配車在庫数", "台", ""),
    ("掲載数", "台", ""),
    ("未掲載数", "台", ""),
    ("キャパ台数", "台", "設定側で管理する場合はそのシート名も記入"),
    ("キャパ使用率", "%", "算出式なら備考に式を記入"),
    ("回転率", "%", "算出式なら備考に式を記入"),
    ("在庫金額", "金額", ""),
    ("在庫金額平均", "金額", "母数は在庫台数"),
]:
    add("販売詳細", "在庫関連", item, unit, item, "高", note)

for item, unit in [
    ("納車MQ", "金額"), ("未納車MQ", "金額"), ("付帯MQ", "金額"), ("KBMQ", "金額"),
    ("当月受注納車Q", "台"), ("当月受注未納車Q", "台"), ("前月受注納車Q", "台"), ("前月受注未納車Q", "台"),
    ("当月即納比率", "%"), ("納車Q前月比率", "%"), ("納車Q当月比率", "%"),
    ("翌月納車Q", "台"), ("翌月納車MQ", "金額")
]:
    add("販売詳細", "受注・納車関連", item, unit, item)

for base in ["クレジット", "保証", "メンテナンス", "クリーニング", "コーティング", "エアコン", "楽々納車", "タイヤ交換"]:
    add("販売詳細", "販売付帯関連", f"{base}件数", "件", f"{base}件数 / {base}Q")
    add("販売詳細", "販売付帯関連", f"{base}MQ", "金額", f"{base}MQ")

# Employees
for item, unit in [
    ("氏名", "文字"), ("営業部名", "文字"), ("拠点名", "文字"), ("役職", "文字"),
    ("ランク", "文字"), ("社歴", "文字"), ("権限", "文字"), ("雇用区分", "文字"),
    ("固定給", "金額"), ("経費", "金額"), ("高速使用料", "金額"), ("MQ", "金額")
]:
    add("社員管理", "社員名簿/組織図/プロフィール", item, unit, item)

# Flash/report/settings
for item, unit in [
    ("速報_営業部", "文字"), ("速報_氏名", "文字"), ("速報_車名", "文字"),
    ("買取速報_媒体", "文字"), ("買取速報_買値", "金額"), ("買取速報_販路区分", "文字"),
    ("販売速報_媒体", "文字"), ("販売速報_粗利", "金額 千円入力"), ("販売速報_備考", "文字"),
    ("速報_査定台数", "件")
]:
    add("速報", "当日速報", item, unit, item, "中")

for item, unit in [
    ("対象月", "年月"), ("日報保存先", "シート/Drive"), ("月締め保存先", "シート/Drive"),
    ("G目標", "金額"), ("MQ目標", "金額"), ("査定Q目標", "件"), ("買取Q目標", "台"),
    ("販売Q目標", "台"), ("査定CVR目標", "%"), ("成約CVR目標", "%"),
    ("権限ロール", "文字"), ("表示範囲", "文字"), ("申請可否", "可否")
]:
    add("設定/日報", "対象月・目標・権限", item, unit, item, "中")

wb = Workbook()
ws = wb.active
ws.title = "全項目_記入用"

header_fill = PatternFill("solid", fgColor="111111")
header_font = Font(color="FFFFFF", bold=True)
sub_fill = PatternFill("solid", fgColor="F2F2F2")
yellow = PatternFill("solid", fgColor="FFF2CC")
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.append(HEADERS)
for row in ROWS:
    ws.append(row)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col in range(7, 12):
        row[col - 1].fill = yellow

widths = [6, 14, 24, 24, 18, 42, 24, 12, 24, 34, 16, 10, 42]
for idx, width in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(idx)].width = width

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

priority_dv = DataValidation(type="list", formula1='"高,中,低"', allow_blank=True)
ws.add_data_validation(priority_dv)
priority_dv.add(f"L2:L{ws.max_row}")

guide = wb.create_sheet("記入ルール")
guide_rows = [
    ["目的", "このブックに、TJ-SalesOSの各表示項目が基幹スプシのどこから来るかを記入してください。"],
    ["記入_参照シート名", "例: 営業部DASH_DB、販売本部DASH_DB、社員名簿リンク、拠点名リンク"],
    ["記入_列番号", "例: AC、FC、DO:ED。単一列ならAC、範囲ならDO:EDのように記入。"],
    ["記入_列名", "1行目ヘッダー名がある場合はその名称。ヘッダー無しなら空欄でOK。"],
    ["記入_集計条件", "例: 営業部名=千葉営業部、拠点名=A列、営業部=V列、対象月=2026/05。"],
    ["記入_例の値", "実データで確認できる値。例: 千葉営業部 受注Q=12、在庫台数=7。"],
    ["金額", "スプシ側が千円単位なら、そのまま千円単位で記入。OS側で円全桁表示にします。"],
    ["優先度", "高は画面表示に必須。中は分析/日報/設定。低は後回し可能。"],
]
for row in guide_rows:
    guide.append(row)
guide.column_dimensions["A"].width = 24
guide.column_dimensions["B"].width = 110
for row in guide.iter_rows():
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
guide["A1"].fill = header_fill
guide["A1"].font = header_font
guide["B1"].fill = header_fill
guide["B1"].font = header_font

summary = wb.create_sheet("メニュー別一覧")
summary.append(["メニュー", "項目数", "記入完了数", "残数"])
menus = []
for row in ROWS:
    if row[1] not in menus:
        menus.append(row[1])
for menu in menus:
    row_num = summary.max_row + 1
    summary.append([
        menu,
        f'=COUNTIF(全項目_記入用!B:B,A{row_num})',
        f'=COUNTIFS(全項目_記入用!B:B,A{row_num},全項目_記入用!G:G,"<>")',
        f'=B{row_num}-C{row_num}',
    ])
for cell in summary[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border
for row in summary.iter_rows(min_row=2):
    for cell in row:
        cell.border = border
summary.column_dimensions["A"].width = 18
summary.column_dimensions["B"].width = 14
summary.column_dimensions["C"].width = 14
summary.column_dimensions["D"].width = 14

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUTPUT)
print(OUTPUT)
