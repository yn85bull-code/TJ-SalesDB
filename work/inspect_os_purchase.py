from openpyxl import load_workbook

p = r"C:\Users\user\Downloads\営業部ダッシュボード (1).xlsx"
wb = load_workbook(p, data_only=True, read_only=True)
ws = wb["OS買取詳細用"]
print("rows", ws.max_row, "cols", ws.max_column)
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
print("HEADERS")
for i, h in enumerate(headers, start=1):
    print(i, h)
print("ROWS")
for r in range(2, min(ws.max_row, 12) + 1):
    vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
    print("ROW", r)
    for i, (h, v) in enumerate(zip(headers, vals), start=1):
        if v not in (None, ""):
            print(i, h, "=", v)
