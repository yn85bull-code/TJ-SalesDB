import json
from pathlib import Path
from openpyxl import load_workbook

DOWNLOADED_SRC = Path("work/master-download.xlsx")
SRC = DOWNLOADED_SRC if DOWNLOADED_SRC.exists() else Path(r"C:\Users\user\Downloads\営業部ダッシュボード (1).xlsx")
SHEET_NAME = "OS買取詳細用"
OUT = Path("outputs/tj-sales-os-lite-demo.html")

KEYS = [
    "営業部名",
    "査定Ｑ", "買取Ｑ", "査定CVR", "成約CVR",
    "AA@", "AAQ", "商品Q", "スクラップQ", "代車Q",
    "AA比率", "商品比率", "スクラップ比率", "代車比率",
    "買取キャンセルQ", "キャンセル比率", "入庫平均日数", "即決契約数", "管理契約数",
    "当月AA出品Q", "当月出品Q", "当月出品比率",
    "AA相違Q", "AA相違比率", "当月AAMQ",
    "AA落札Q", "AA未処理Q", "AA流札Q", "AA取消Q", "AA翌月Q",
    "AA赤字金額", "AA赤字Q", "AA赤字比率",
]

wb = load_workbook(SRC, data_only=True)
ws = wb[SHEET_NAME]
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
rows = []
for r in range(2, ws.max_row + 1):
    obj = {}
    empty = True
    for c, h in enumerate(headers, start=1):
        if not h:
            continue
        v = ws.cell(r, c).value
        if v not in (None, ""):
            empty = False
        obj[str(h)] = v
    if not empty and obj.get("営業部名"):
        slim = {k: obj.get(k, "") for k in KEYS}
        slim["当月出品比率"] = ws.cell(r, 18).value
        slim["入庫平均日数"] = ws.cell(r, 19).value
        slim["査定CVR"] = ws.cell(r, 132).value
        slim["即決契約数"] = ws.cell(r, 154).value
        slim["管理契約数"] = ws.cell(r, 155).value
        rows.append(slim)

mock = json.dumps(rows, ensure_ascii=False)

html = f"""<!doctype html>
<html lang="ja">
<head>
  <base target="_top">
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>TJ-SalesOS Lite Demo</title>
  <style>
    :root {{
      --bg:#f2f2f3;
      --panel:#fff;
      --text:#080808;
      --muted:#666;
      --line:#101010;
      --soft-line:#dedede;
      --red:#c90016;
      --black:#080808;
      --white:#fff;
      --cyan:#10aee8;
      --hero-grad:linear-gradient(135deg,#fff 0%,#fff1f3 58%,#ffcbd2 100%);
      --dark-grad:linear-gradient(135deg,#050505 0%,#151515 60%,#c90016 61%,#c90016 100%);
    }}
    * {{ box-sizing:border-box; }}
    body {{
      margin:0;
      background:var(--bg);
      color:var(--text);
      font-family:"Noto Sans JP","Yu Gothic","Meiryo",system-ui,sans-serif;
      letter-spacing:0;
    }}
    .app-shell {{ display:grid; grid-template-columns:240px 1fr; min-height:100vh; }}
    .sidebar {{
      background:var(--black);
      color:var(--white);
      border-right:5px solid var(--red);
      padding:22px 16px;
      position:sticky;
      top:0;
      height:100vh;
    }}
    .sidebar-logo {{ font-size:26px; font-weight:1000; margin-bottom:6px; }}
    .sidebar-sub {{ font-size:11px; color:#d7d7d7; font-weight:900; margin-bottom:22px; }}
    .dept-select {{
      width:100%;
      height:42px;
      border:2px solid #fff;
      border-radius:6px;
      background:#fff;
      color:#111;
      padding:0 10px;
      font-weight:1000;
      margin-bottom:16px;
    }}
    .side-menu {{ display:grid; gap:8px; }}
    .side-menu button {{
      display:flex;
      align-items:center;
      gap:10px;
      width:100%;
      height:48px;
      border:2px solid #fff;
      background:transparent;
      color:#fff;
      text-align:left;
      padding:0 14px;
      border-radius:6px;
      font-weight:1000;
    }}
    .side-menu button.active {{ background:var(--red); border-color:var(--red); }}
    .main-area {{ min-width:0; }}
    .topbar {{
      position:sticky;
      top:0;
      z-index:10;
      height:72px;
      background:var(--black);
      color:var(--white);
      border-bottom:4px solid var(--red);
      display:flex;
      align-items:center;
      justify-content:space-between;
      padding:0 22px;
    }}
    .brand-block {{ display:flex; align-items:baseline; gap:14px; }}
    .brand {{ font-size:24px; font-weight:1000; }}
    .sub-brand {{ color:#d7d7d7; font-size:12px; font-weight:900; }}
    .toolbar {{ display:flex; gap:10px; align-items:center; }}
    .toolbar button {{
      height:40px;
      border:2px solid #fff;
      border-radius:6px;
      background:var(--red);
      color:#fff;
      padding:0 14px;
      font-weight:1000;
    }}
    .content {{ padding:22px; }}
    .status {{ margin:0 0 14px; color:#555; font-size:13px; font-weight:900; }}
    .page-heading {{
      display:flex;
      align-items:center;
      justify-content:space-between;
      gap:12px;
      margin:0 0 14px;
      padding:16px 20px;
      border:3px solid var(--line);
      border-left:8px solid var(--red);
      background:#fff;
      font-size:22px;
      font-weight:1000;
      box-shadow:0 5px 0 rgba(0,0,0,.08);
    }}
    .panel {{
      background:var(--panel);
      border:3px solid var(--line);
      border-radius:8px;
      padding:16px;
      margin-bottom:18px;
      box-shadow:0 7px 0 rgba(0,0,0,.08);
    }}
    .section-title {{
      margin:0 0 14px;
      padding-bottom:10px;
      border-bottom:3px solid var(--line);
      color:var(--red);
      font-size:16px;
      font-weight:1000;
    }}
    .grid {{ display:grid; gap:12px; }}
    .top-kpi {{ grid-template-columns:repeat(2,minmax(0,1fr)); }}
    .two-col {{ grid-template-columns:1.08fr .92fr; }}
    .mini-2 {{ grid-template-columns:repeat(2,minmax(0,1fr)); }}
    .mini-3 {{ grid-template-columns:repeat(3,minmax(0,1fr)); }}
    .mini-4 {{ grid-template-columns:repeat(4,minmax(0,1fr)); }}
    .kpi-card {{
      position:relative;
      overflow:hidden;
      min-height:154px;
      border:4px solid var(--line);
      border-radius:8px;
      background:var(--hero-grad);
      padding:18px;
      box-shadow:0 8px 0 #111;
      isolation:isolate;
    }}
    .kpi-card::before {{
      content:"";
      position:absolute;
      inset:-40%;
      z-index:0;
      background:
        radial-gradient(circle at 18% 24%, rgba(255,255,255,.95) 0 2px, transparent 3px),
        radial-gradient(circle at 72% 34%, rgba(255,255,255,.9) 0 2px, transparent 3px),
        linear-gradient(115deg, transparent 34%, rgba(255,255,255,.88) 46%, transparent 58%);
      opacity:.7;
      transform:translateX(-34%) rotate(8deg);
      animation:shine 2.6s linear infinite;
      pointer-events:none;
    }}
    .kpi-card > * {{ position:relative; z-index:1; }}
    .flag {{
      display:inline-flex;
      align-items:center;
      justify-content:center;
      min-width:116px;
      height:34px;
      border:3px solid #000;
      border-radius:5px;
      background:linear-gradient(180deg,#e0001b,#9b0014 60%,#65000b);
      color:#fff;
      padding:0 14px;
      font-size:15px;
      font-weight:1000;
      box-shadow:4px 4px 0 #111;
    }}
    .kpi-main {{ margin-top:22px; display:grid; grid-template-columns:minmax(0,1fr) minmax(260px,42%); align-items:end; gap:18px; }}
    .kpi-value {{ font-size:60px; line-height:.9; font-weight:1000; letter-spacing:-1px; white-space:nowrap; }}
    .unit {{ font-size:.55em; margin-left:3px; }}
    .rate-box {{
      width:100%;
      min-height:118px;
      border:3px solid var(--red);
      border-radius:6px;
      background:linear-gradient(135deg,#ffffff 0%,#f7fbff 58%,#eef5ff 100%);
      color:#071d3a;
      display:flex;
      flex-direction:column;
      align-items:center;
      justify-content:center;
      text-align:center;
      box-shadow:4px 4px 0 rgba(0,0,0,.35), inset 0 0 0 2px #071d3a;
    }}
    .rate-box span {{ font-size:18px; font-weight:1000; margin-bottom:10px; }}
    .rate-box b {{ font-size:54px; line-height:1; font-weight:1000; }}
    .rate-box b .js-count {{ font-size:54px; line-height:1; }}
    .box {{
      min-width:0;
      border:2px solid var(--line);
      border-radius:7px;
      background:#fff;
      padding:12px;
    }}
    .box.tint {{ background:var(--hero-grad); }}
    .label {{ color:#111; font-size:12px; font-weight:1000; margin-bottom:8px; }}
    .value {{ font-size:26px; line-height:1; font-weight:1000; white-space:nowrap; }}
    .negative {{ color:var(--red); }}
    .bar-meta {{ display:flex; justify-content:space-between; gap:8px; margin:9px 0 4px; font-size:11px; font-weight:1000; }}
    .track {{ height:14px; border:3px solid #000; border-radius:12px; background:#fff; overflow:hidden; }}
    .fill {{ height:100%; width:0; background:linear-gradient(90deg,#c90016,#65000b 65%,#111); border-right:3px solid #000; transition:width 1.1s cubic-bezier(.16,1,.3,1); }}
    .purchase-grid {{ grid-template-columns:minmax(0,1.12fr) minmax(0,.88fr); align-items:start; }}
    .tight-box {{ padding:10px; }}
    .route-grid {{ display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:8px; }}
    .route-item {{
      min-width:0;
      border:2px solid #111;
      border-radius:6px;
      padding:9px 10px;
      background:#fff;
    }}
    .route-item .label {{ margin-bottom:6px; }}
    .route-item .value {{ font-size:22px; }}
    .route-item .track {{ height:11px; border-width:2px; }}
    .route-item .bar-meta {{ margin-top:6px; font-size:10px; }}
    .contract-grid {{ display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:10px; }}
    .contract-item {{
      border:2px solid #111;
      border-radius:7px;
      padding:10px;
      background:linear-gradient(135deg,#fff 0%,#fff7f8 55%,#ffdbe2 100%);
    }}
    .contract-item .value {{ font-size:22px; }}
    .inner-block {{
      border:2px solid #111;
      border-radius:7px;
      padding:10px;
      background:linear-gradient(135deg,#fff 0%,#fff8f9 62%,#ffe3e8 100%);
      margin-top:10px;
    }}
    .inner-block:first-of-type {{ margin-top:0; }}
    .stackbar {{ height:24px; display:flex; overflow:hidden; border:3px solid #000; border-radius:14px; background:#fff; }}
    .seg {{ width:0; height:100%; border-right:2px solid #000; transition:width 1.1s cubic-bezier(.16,1,.3,1); }}
    .red {{ background:linear-gradient(180deg,#e0001b,#9b0014); }}
    .black {{ background:linear-gradient(180deg,#333,#050505); }}
    .blue {{ background:linear-gradient(180deg,#1fb6ff,#0077c8); }}
    .yellow {{ background:linear-gradient(180deg,#ffd247,#ffb000); }}
    .legend {{ display:grid; grid-template-columns:1fr 1fr; gap:7px 14px; margin-top:10px; font-size:11px; font-weight:950; }}
    .dot {{ display:inline-block; width:9px; height:9px; border-radius:50%; border:2px solid #111; margin-right:5px; }}
    .dot.red {{ background:#c90016; }}
    .dot.black {{ background:#111; }}
    .dot.blue {{ background:#0077c8; }}
    .dot.yellow {{ background:#ffb000; }}
    .block-title {{
      display:inline-flex;
      align-items:center;
      min-height:28px;
      margin:0 0 10px;
      padding:0 12px;
      border:2px solid #73000e;
      border-radius:5px;
      background:linear-gradient(180deg,#d6001c,#8b0011);
      color:#fff;
      font-size:13px;
      font-weight:1000;
      box-shadow:2px 2px 0 #111;
    }}
    .aa-main {{
      min-height:300px;
      display:grid;
      gap:10px;
      align-content:stretch;
      border:4px solid var(--line);
      border-radius:8px;
      background:var(--hero-grad);
      padding:18px;
      box-shadow:0 8px 0 #111;
      position:relative;
      overflow:hidden;
      isolation:isolate;
    }}
    .aa-main::before {{
      content:"";
      position:absolute;
      inset:-40%;
      z-index:0;
      background:linear-gradient(115deg, transparent 34%, rgba(255,255,255,.88) 46%, transparent 58%);
      opacity:.7;
      animation:shine 2.6s linear infinite;
    }}
    .aa-main > * {{ position:relative; z-index:1; }}
    .aa-big {{ font-size:76px; font-weight:1000; line-height:.9; }}
    .counting {{ animation:numberPop .22s ease-out; }}
    @keyframes shine {{
      0% {{ transform:translateX(-42%) rotate(8deg); opacity:.28; }}
      38% {{ opacity:.82; }}
      100% {{ transform:translateX(38%) rotate(8deg); opacity:.28; }}
    }}
    @keyframes numberPop {{
      0% {{ transform:scale(.94); filter:brightness(1.2); }}
      100% {{ transform:scale(1); filter:brightness(1); }}
    }}
    @media (max-width:900px) {{
      .app-shell {{ grid-template-columns:1fr; }}
      .sidebar {{ position:relative; height:auto; }}
      .top-kpi,.two-col,.purchase-grid,.mini-2,.mini-3,.mini-4 {{ grid-template-columns:1fr; }}
      .kpi-main {{ grid-template-columns:1fr; }}
      .kpi-value {{ font-size:48px; }}
      .route-grid,.contract-grid {{ grid-template-columns:repeat(2,minmax(0,1fr)); }}
      .route-item .value,.contract-item .value {{ font-size:20px; }}
    }}
  </style>
</head>
<body>
  <div class="app-shell">
    <aside class="sidebar">
      <div class="sidebar-logo">TJ-SalesOS</div>
      <div class="sidebar-sub">Sales Operating System</div>
      <select id="deptSelect" class="dept-select"></select>
      <nav class="side-menu">
        <button type="button">⌂ HOME</button>
        <button type="button">♡ 実績管理</button>
        <button type="button" class="active">♤ 買取詳細</button>
        <button type="button">♢ 販売詳細</button>
        <button type="button">♧ 社員管理</button>
      </nav>
    </aside>

    <main class="main-area">
      <header class="topbar">
        <div class="brand-block">
          <div class="brand">TJ-SalesOS</div>
          <div class="sub-brand">OS買取詳細用 / Lite Data</div>
        </div>
        <div class="toolbar">
          <button type="button" onclick="reloadLite()">再読み込み</button>
        </div>
      </header>
      <section class="content">
        <p class="status" id="status">買取詳細 / - / デモ表示</p>
        <h1 class="page-heading"><span>買取詳細</span><span id="titleDept">-</span></h1>
        <div id="content"></div>
      </section>
    </main>
  </div>

  <script>
    const MOCK_ROWS = {mock};
    let rows = [];
    let current = null;

    const moneyKKeys = new Set(['AA@','当月AAMQ','AA赤字金額']);
    const percentKeys = new Set(['査定CVR','成約CVR','AA比率','商品比率','スクラップ比率','代車比率','当月出品比率','AA相違比率','キャンセル比率','AA赤字比率']);

    function runGas(name, args) {{
      return new Promise((resolve, reject) => {{
        if (!window.google || !google.script || !google.script.run) return resolve(MOCK_ROWS);
        google.script.run.withSuccessHandler(resolve).withFailureHandler(reject)[name].apply(google.script.run, args || []);
      }});
    }}
    function esc(v) {{ return String(v ?? '').replace(/[&<>"']/g, s => ({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#039;'}}[s])); }}
    function num(v) {{ const n = Number(String(v ?? '').replace(/[¥￥,%台件日\\s]/g,'')); return isNaN(n) ? 0 : n; }}
    function pct(v) {{ const n = num(v); return Math.abs(n) <= 1 ? n * 100 : n; }}
    function rate(v,total) {{ const t = num(total); return t ? num(v) / t * 100 : 0; }}
    function clamp(n) {{ return Math.max(0, Math.min(100, n || 0)); }}
    function dataType(key) {{ if (percentKeys.has(key)) return 'percent'; if (moneyKKeys.has(key)) return 'moneyK'; if (key === '入庫平均日数') return 'days'; return 'number'; }}
    function formatByType(type, value, suffix='') {{
      const n = Number(value) || 0;
      if (type === 'percent') return `${{(Math.abs(n) <= 1 ? n * 100 : n).toFixed(1)}}%`;
      if (type === 'percentPlain') return `${{n.toFixed(1)}}%`;
      if (type === 'moneyK') return `${{n < 0 ? '-' : ''}}¥${{Math.abs(Math.round(n * 1000)).toLocaleString('ja-JP')}}${{suffix}}`;
      if (type === 'days') return `${{n.toFixed(1).replace(/\\.0$/,'')}}${{suffix}}`;
      return `${{Math.round(n).toLocaleString('ja-JP')}}${{suffix}}`;
    }}
    function valueHtml(key, value, suffix='') {{
      return `<span class="js-count" data-count-key="${{esc(key)}}" data-count-value="${{esc(num(value))}}" data-count-type="${{esc(dataType(key))}}" data-count-suffix="${{esc(suffix)}}">${{esc(formatByType(dataType(key), num(value), suffix))}}</span>`;
    }}
    function valClass(v) {{ return num(v) < 0 ? 'negative' : ''; }}
    function kpi(label,key,unit,rateLabel,rateKey) {{
      return `<article class="kpi-card">
        <span class="flag">${{esc(label)}}</span>
        <div class="kpi-main">
          <div class="kpi-value">${{valueHtml(key,current?.[key])}}<span class="unit">${{esc(unit)}}</span></div>
          <div class="rate-box"><span>${{esc(rateLabel)}}</span><b>${{valueHtml(rateKey,current?.[rateKey])}}</b></div>
        </div>
      </article>`;
    }}
    function metric(label,key,suffix='',tint=false) {{
      const v = current?.[key];
      return `<article class="box ${{tint ? 'tint' : ''}}">
        <div class="label">${{esc(label)}}</div>
        <div class="value ${{valClass(v)}}">${{valueHtml(key,v,suffix)}}</div>
      </article>`;
    }}
    function ratioMetric(label,key,totalKey,suffix='台') {{
      const v = current?.[key];
      const r = rate(v,current?.[totalKey]);
      return `<article class="box">
        <div class="label">${{esc(label)}}</div>
        <div class="value">${{valueHtml(key,v,suffix)}} / <span class="js-count" data-count-type="percentPlain" data-count-value="${{esc(r)}}">${{r.toFixed(1)}}%</span></div>
        <div class="bar-meta"><span>${{esc(label)}}比率</span><span>${{r.toFixed(1)}}%</span></div>
        <div class="track"><div class="fill js-bar" data-target-width="${{clamp(r)}}"></div></div>
      </article>`;
    }}
    function ratioMetricFromRate(label,key,rateKey,suffix='台') {{
      const v = current?.[key];
      const r = pct(current?.[rateKey]);
      return `<article class="box">
        <div class="label">${{esc(label)}}</div>
        <div class="value">${{valueHtml(key,v,suffix)}} / <span class="js-count" data-count-type="percentPlain" data-count-value="${{esc(r)}}">${{r.toFixed(1)}}%</span></div>
        <div class="bar-meta"><span>${{esc(rateKey)}}</span><span>${{r.toFixed(1)}}%</span></div>
        <div class="track"><div class="fill js-bar" data-target-width="${{clamp(r)}}"></div></div>
      </article>`;
    }}
    function ratioTile(label,key,totalKey,suffix='台') {{
      const v = current?.[key];
      const r = rate(v,current?.[totalKey]);
      return `<div class="route-item">
        <div class="label">${{esc(label)}}</div>
        <div class="value">${{valueHtml(key,v,suffix)}} / <span class="js-count" data-count-type="percentPlain" data-count-value="${{esc(r)}}">${{r.toFixed(1)}}%</span></div>
        <div class="bar-meta"><span>${{esc(label)}}比率</span><span>${{r.toFixed(1)}}%</span></div>
        <div class="track"><div class="fill js-bar" data-target-width="${{clamp(r)}}"></div></div>
      </div>`;
    }}
    function contractCards() {{
      const items = [
        {{ label:'即決契約', key:'即決契約数' }},
        {{ label:'管理契約', key:'管理契約数' }}
      ];
      return items.map(it => {{
        const v = current?.[it.key];
        const r = rate(v,current?.['買取Ｑ']);
        return `<div class="contract-item">
          <div class="label">${{esc(it.label)}}</div>
          <div class="value">${{valueHtml(it.key,v,'件')}} / <span class="js-count" data-count-type="percentPlain" data-count-value="${{esc(r)}}">${{r.toFixed(1)}}%</span></div>
          <div class="bar-meta"><span>${{esc(it.label)}}比率</span><span>${{r.toFixed(1)}}%</span></div>
          <div class="track"><div class="fill js-bar" data-target-width="${{clamp(r)}}"></div></div>
        </div>`;
      }}).join('');
    }}
    function contractBox() {{
      const cards = contractCards();
      return `<article class="box"><h3 class="block-title">契約内訳</h3><div class="contract-grid">${{cards}}</div></article>`;
    }}
    function contractManagementBox() {{
      const cancelValue = current?.['買取キャンセルQ'];
      const cancelRate = rate(cancelValue,current?.['買取Ｑ']);
      return `<article class="box">
        <h3 class="block-title">契約・入庫管理</h3>
        <div class="inner-block"><div class="contract-grid">${{contractCards()}}</div></div>
        <div class="inner-block">
          <div class="label">入庫平均日数</div>
          <div class="value">${{valueHtml('入庫平均日数',current?.['入庫平均日数'],'日')}}</div>
        </div>
        <div class="inner-block">
          <div class="label">買取キャンセル</div>
          <div class="value">${{valueHtml('買取キャンセルQ',cancelValue,'台')}} / <span class="js-count" data-count-type="percentPlain" data-count-value="${{esc(cancelRate)}}">${{cancelRate.toFixed(1)}}%</span></div>
          <div class="bar-meta"><span>買取キャンセル比率</span><span>${{cancelRate.toFixed(1)}}%</span></div>
          <div class="track"><div class="fill js-bar" data-target-width="${{clamp(cancelRate)}}"></div></div>
        </div>
      </article>`;
    }}
    function stack(title,items,totalKey) {{
      const colors = ['red','black','blue','yellow'];
      const total = Math.max(1,num(current?.[totalKey]) || items.reduce((s,it)=>s+num(current?.[it.key]),0));
      const segs = items.map((it,i)=>`<span class="seg ${{colors[i%colors.length]}} js-seg" data-target-width="${{clamp(num(current?.[it.key])/total*100)}}"></span>`).join('');
      const legend = items.map((it,i)=>`<span><i class="dot ${{colors[i%colors.length]}}"></i>${{esc(it.label)}} ${{Math.round(num(current?.[it.key])).toLocaleString('ja-JP')}}台 / ${{(num(current?.[it.key])/total*100).toFixed(1)}}%</span>`).join('');
      return `<article class="box"><h3 class="block-title">${{esc(title)}}</h3><div class="stackbar">${{segs}}</div><div class="legend">${{legend}}</div></article>`;
    }}
    function render() {{
      const dept = current?.['営業部名'] || '-';
      document.getElementById('titleDept').textContent = dept;
      document.getElementById('status').textContent = `買取詳細 / ${{dept}} / OS買取詳細用`;
      document.getElementById('content').innerHTML = `
        <section class="panel">
          <h2 class="section-title">買取実績詳細</h2>
          <div class="grid top-kpi">
            ${{kpi('査定数','査定Ｑ','件','査定CVR','査定CVR')}}
            ${{kpi('買取台数','買取Ｑ','台','成約CVR','成約CVR')}}
          </div>
          <div class="grid purchase-grid" style="margin-top:14px;">
            <div class="grid">
              <article class="box tight-box">
                <h3 class="block-title">販路内訳</h3>
                <div class="route-grid">
                  ${{ratioTile('AA','AAQ','買取Ｑ')}}
                  ${{ratioTile('商品','商品Q','買取Ｑ')}}
                  ${{ratioTile('スクラップ','スクラップQ','買取Ｑ')}}
                  ${{ratioTile('代車','代車Q','買取Ｑ')}}
                </div>
              </article>
              ${{stack('販路構成',[
                {{label:'AA',key:'AAQ'}},
                {{label:'商品',key:'商品Q'}},
                {{label:'スクラップ',key:'スクラップQ'}},
                {{label:'代車',key:'代車Q'}}
              ],'買取Ｑ')}}
            </div>
            <div class="grid">
              ${{contractManagementBox()}}
            </div>
          </div>
        </section>

        <section class="panel">
          <h2 class="section-title" style="color:#111;">AA出品詳細</h2>
          <div class="grid two-col">
            <article class="aa-main">
              <div>
                <div class="label">出品Q</div>
                <div class="aa-big">${{valueHtml('当月AA出品Q',current?.['当月AA出品Q'])}}<span class="unit">台</span></div>
              </div>
              <div class="grid mini-2">
                ${{metric('AAMQ','当月AAMQ')}}
                ${{metric('AA@','AA@')}}
                ${{metric('処理予定台数','AA未処理Q','台')}}
                ${{metric('処理予定MQ','当月AAMQ')}}
              </div>
            </article>
            <div class="grid">
              <div class="grid mini-3">
                ${{metric('落札Q','AA落札Q','台')}}
                ${{metric('流札Q','AA流札Q','台')}}
                ${{metric('取消Q','AA取消Q','台')}}
              </div>
              ${{stack('出品Q内訳',[
                {{label:'落札',key:'AA落札Q'}},
                {{label:'流札',key:'AA流札Q'}},
                {{label:'取消',key:'AA取消Q'}}
              ],'当月AA出品Q')}}
              <div class="grid mini-2">
                ${{metric('赤字Q比率','AA赤字比率')}}
                ${{metric('赤字金額','AA赤字金額')}}
              </div>
            </div>
          </div>
          <div class="grid mini-4" style="margin-top:12px;">
            ${{ratioMetricFromRate('当月買取出品Q','当月出品Q','当月出品比率')}}
            ${{ratioMetric('AA相違','AA相違Q','当月AA出品Q')}}
            ${{ratioMetric('落札率','AA落札Q','当月AA出品Q')}}
            ${{metric('持越し台数','AA翌月Q','台',true)}}
          </div>
        </section>`;
      runAnimations();
    }}
    function renderCount(type,value,key='',suffix='') {{
      return formatByType(type, value, suffix);
    }}
    function runAnimations() {{
      const root = document.getElementById('content');
      root.querySelectorAll('.js-bar,.js-seg').forEach(el=>{{
        el.style.width='0%';
        const target = Number(el.dataset.targetWidth || 0);
        requestAnimationFrame(()=>requestAnimationFrame(()=>{{ el.style.width = `${{target}}%`; }}));
      }});
      root.querySelectorAll('.js-count').forEach(el=>{{
        const target = Number(el.dataset.countValue || 0);
        const type = el.dataset.countType || 'number';
        const suffix = el.dataset.countSuffix || '';
        const start = performance.now();
        const duration = type === 'percent' || type === 'percentPlain' ? 850 : 1050;
        el.classList.add('counting');
        function tick(now) {{
          const t = Math.min(1,(now-start)/duration);
          const eased = 1 - Math.pow(1-t,3);
          const value = target * eased;
          el.textContent = renderCount(type,value,'',suffix);
          if (t < 1) requestAnimationFrame(tick);
          else {{
            el.textContent = renderCount(type,target,'',suffix);
            setTimeout(()=>el.classList.remove('counting'),80);
          }}
        }}
        requestAnimationFrame(tick);
      }});
    }}
    async function reloadLite() {{
      const previousDept = document.getElementById('deptSelect')?.value || '';
      rows = await runGas('getLiteDashboardRows');
      rows = Array.isArray(rows) ? rows.filter(r => r && r['営業部名']) : MOCK_ROWS;
      const select = document.getElementById('deptSelect');
      select.innerHTML = rows.map(r => `<option>${{esc(r['営業部名'])}}</option>`).join('');
      current = rows.find(r => r['営業部名'] === previousDept) || rows.find(r => r['営業部名'] === '千葉営業部') || rows[0];
      select.value = current?.['営業部名'] || '';
      select.onchange = () => {{ current = rows.find(r => r['営業部名'] === select.value) || rows[0]; render(); }};
      render();
    }}
    reloadLite();
  </script>
</body>
</html>
"""

OUT.write_text(html, encoding="utf-8")
print(OUT)
