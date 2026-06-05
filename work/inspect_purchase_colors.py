from openpyxl import load_workbook

p = r"C:\Users\user\Downloads\営業部ダッシュボード (1).xlsx"
wb = load_workbook(p, data_only=True)
ws = wb["OS買取詳細用"]
for c in range(1, min(ws.max_column, 120) + 1):
    cell = ws.cell(1, c)
    color = cell.fill.fgColor.rgb or str(cell.fill.fgColor.indexed) or cell.fill.fgColor.type
    print(c, cell.value, color)
