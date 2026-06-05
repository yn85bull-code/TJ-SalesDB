from openpyxl import load_workbook

p = r"C:\Users\user\Downloads\TJ-SalesOS_参照マッピング記入表.xlsx"
wb = load_workbook(p, data_only=True)
print(wb.sheetnames)
ws = wb["全項目_記入用"]
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
print(headers)
filled = []
for r in range(2, ws.max_row + 1):
    row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
    if any(row.get(k) for k in ["記入_参照シート名", "記入_列番号", "記入_列名", "記入_集計条件", "記入_例の値"]):
        filled.append(row)
print("filled", len(filled), "of", ws.max_row - 1)
for row in filled:
    print(
        row["No"],
        row["メニュー"],
        row["セクション"],
        row["OS表示項目"],
        row.get("記入_参照シート名"),
        row.get("記入_列番号"),
        row.get("記入_列名"),
        row.get("記入_集計条件"),
        row.get("記入_例の値"),
    )
