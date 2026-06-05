from openpyxl import load_workbook

wb = load_workbook("work/master-download.xlsx", data_only=True)
targets = ["OS実績管理用", "OS買取詳細用", "OS販売詳細用", "営業部DASH_DB", "営業部DASH_目標"]

for ws in wb.worksheets:
    if ws.title not in targets:
        continue
    print(f"\n## {ws.title} rows={ws.max_row} cols={ws.max_column}")
    headers = [ws.cell(1, c).value for c in range(1, min(ws.max_column, 220) + 1)]
    for i, h in enumerate(headers, start=1):
        if h:
            print(i, h)
