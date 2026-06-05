import json
from datetime import date, datetime
from pathlib import Path
from openpyxl import load_workbook

SRC = Path("work/master-download.xlsx")
OUT = Path("outputs/tj-sales-os-onepage.html")

wb = load_workbook(SRC, data_only=True)

def sheet_objects(sheet_name):
    ws = wb[sheet_name]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        obj = {}
        has_value = False
        for c, h in enumerate(headers, start=1):
            if not h:
                continue
            v = ws.cell(r, c).value
            if isinstance(v, (datetime, date)):
                v = v.strftime("%Y/%m/%d %H:%M") if isinstance(v, datetime) else v.strftime("%Y/%m/%d")
            if v not in (None, ""):
                has_value = True
            obj[str(h)] = v
        if has_value and obj.get("営業部名"):
            rows.append(obj)
    return rows

payload = {
    "osRows": sheet_objects("OS実績管理用"),
    "purchaseRows": sheet_objects("OS買取詳細用"),
    "salesRows": sheet_objects("OS販売詳細用"),
    "dashRows": sheet_objects("営業部DASH_DB"),
}

html = r"""<!doctype html>
<html lang="ja">
<head>
  <base target="_top">
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>TJ-SalesOS 営業部サマリー</title>
  <style>
    :root {
      --bg:#f3f3f4;
      --panel:#fff;
      --ink:#080808;
      --muted:#5d626b;
      --red:#d7001d;
      --red-dark:#850012;
      --navy:#071d3a;
      --line:#0b0b0c;
      --blue:#0877c7;
      --yellow:#ffbd18;
      --green:#0c8f5a;
    }
    * { box-sizing:border-box; }
    body {
      margin:0;
      background:var(--bg);
      color:var(--ink);
      font-family:Arial,"Yu Gothic","Meiryo",sans-serif;
      font-weight:800;
    }
    .one-page-dashboard { min-width:1180px; }
    .app-header {
      position:sticky;
      top:0;
      z-index:20;
      display:grid;
      grid-template-columns:290px 1fr auto;
      gap:18px;
      align-items:center;
      padding:18px 24px;
      background:#050505;
      color:#fff;
      border-bottom:5px solid var(--red);
    }
    .brand-title { font-size:28px; font-weight:1000; letter-spacing:.2px; }
    .brand-sub { margin-top:3px; font-size:12px; color:#cfd3da; }
    .control-row { display:flex; align-items:center; gap:12px; flex-wrap:wrap; }
    select,button {
      height:42px;
      border:2px solid #fff;
      border-radius:7px;
      font-weight:1000;
      font-size:14px;
    }
    select { min-width:230px; padding:0 12px; color:#111; background:#fff; }
    button { padding:0 16px; color:#fff; background:linear-gradient(180deg,#e10020,#990012); cursor:pointer; }
    .header-meta { display:flex; gap:10px; justify-content:flex-end; color:#fff; font-size:12px; }
    .meta-pill { border:1px solid rgba(255,255,255,.45); border-radius:999px; padding:8px 10px; background:rgba(255,255,255,.08); }
    .page { padding:22px; }
    .summary-hero {
      display:grid;
      grid-template-columns:1.1fr .9fr;
      gap:16px;
      background:#fff;
      border:3px solid var(--line);
      border-left:10px solid var(--red);
      border-radius:8px;
      padding:18px;
      box-shadow:0 7px 0 rgba(0,0,0,.10);
    }
    .dept-name { font-size:42px; font-weight:1000; line-height:1; }
    .hero-note { color:var(--muted); margin-top:8px; font-size:13px; }
    .staff-grid { display:grid; grid-template-columns:repeat(5,1fr); gap:10px; }
    .staff-card,.kpi-card,.ratio-card,.mq-card,.progress-card,.detail-card {
      min-width:0;
      border:2px solid var(--line);
      border-radius:8px;
      background:#fff;
      padding:12px;
    }
    .staff-card { background:linear-gradient(135deg,#fff,#fff5f6); }
    .label { color:#1c1d20; font-size:12px; font-weight:1000; margin-bottom:7px; }
    .value { font-size:26px; line-height:1; font-weight:1000; white-space:nowrap; }
    .unit { font-size:.56em; margin-left:2px; }
    .negative { color:var(--red); }
    .major-section {
      margin-top:18px;
      padding:16px;
      border:3px solid var(--line);
      border-radius:8px;
      background:var(--panel);
      box-shadow:0 6px 0 rgba(0,0,0,.08);
    }
    .section-head {
      display:flex;
      justify-content:space-between;
      align-items:center;
      gap:12px;
      margin-bottom:12px;
      padding-bottom:10px;
      border-bottom:3px solid var(--line);
    }
    .section-title { font-size:18px; font-weight:1000; color:var(--red); }
    .section-tag { color:#fff; background:#0a0a0b; border-radius:999px; padding:6px 10px; font-size:11px; }
    .core-kpi-grid { display:grid; grid-template-columns:1.25fr 1fr 1fr; gap:14px; }
    .kpi-card.big {
      min-height:145px;
      display:flex;
      flex-direction:column;
      justify-content:space-between;
      background:linear-gradient(135deg,#fff 0%,#fff6f7 50%,#ffdbe2 100%);
      border-width:4px;
      box-shadow:0 8px 0 #111;
    }
    .kpi-card.big .value { font-size:46px; }
    .section-grid-3 { display:grid; grid-template-columns:repeat(3,minmax(0,1fr)); gap:12px; }
    .section-grid-4 { display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:12px; }
    .section-grid-2 { display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:12px; }
    .sub-section { margin-top:12px; }
    .ratio-card .value { font-size:23px; }
    .bar-meta { display:flex; justify-content:space-between; gap:8px; margin:9px 0 5px; font-size:11px; }
    .progress-bar { height:15px; overflow:hidden; border:3px solid #000; border-radius:999px; background:#fff; }
    .progress-fill { height:100%; width:0; background:linear-gradient(90deg,var(--red),#1a1a1a); transition:width 1.05s cubic-bezier(.16,1,.3,1); }
    .progress-fill.blue { background:linear-gradient(90deg,#0a7bd3,#071d3a); }
    .progress-fill.green { background:linear-gradient(90deg,#00a060,#064b30); }
    .progress-fill.warn { background:linear-gradient(90deg,#ffcd2d,#ff8c00); }
    .progress-fill.bad { background:linear-gradient(90deg,#e50020,#6d000d); }
    .stackbar { display:flex; height:26px; border:3px solid #000; border-radius:999px; overflow:hidden; background:#fff; }
    .seg { width:0; border-right:2px solid #000; transition:width 1.05s cubic-bezier(.16,1,.3,1); }
    .seg.red { background:linear-gradient(180deg,#e0001b,#8d0011); }
    .seg.black { background:linear-gradient(180deg,#222,#050505); }
    .seg.blue { background:linear-gradient(180deg,#1da7ff,#0069b2); }
    .seg.yellow { background:linear-gradient(180deg,#ffd54f,#ffb000); }
    .legend { display:grid; grid-template-columns:repeat(4,1fr); gap:8px; margin-top:8px; font-size:11px; }
    .dot { display:inline-block; width:10px; height:10px; border:2px solid #000; border-radius:50%; margin-right:5px; }
    .dot.red { background:#d7001d; } .dot.black { background:#111; } .dot.blue { background:#0877c7; } .dot.yellow { background:#ffbd18; }
    .warn-card { background:#fff7df; }
    .bad-card { background:#fff0f2; color:#a40016; }
    .money { letter-spacing:.2px; }
    .footer-note { margin:18px 0 0; color:#60656f; font-size:12px; }
    @media (max-width:1200px) {
      .one-page-dashboard { min-width:0; }
      .app-header,.summary-hero,.core-kpi-grid,.section-grid-2,.section-grid-3,.section-grid-4 { grid-template-columns:1fr; }
      .staff-grid { grid-template-columns:repeat(2,1fr); }
    }
  </style>
</head>
<body>
  <div class="one-page-dashboard">
    <header class="app-header">
      <div>
        <div class="brand-title">TJ-SalesOS</div>
        <div class="brand-sub">営業部サマリー / One Page View</div>
      </div>
      <div class="control-row">
        <select id="deptSelect" onchange="loadSelectedDept()"></select>
        <button type="button" onclick="refreshDashboard()">DASH更新</button>
        <button type="button" onclick="reloadOnePage()">再読み込み</button>
      </div>
      <div class="header-meta">
        <span class="meta-pill" id="targetMonth">対象月 -</span>
        <span class="meta-pill" id="updatedAt">更新 -</span>
      </div>
    </header>
    <main class="page" id="app"></main>
  </div>

  <script>
    const EMBEDDED = __DATA__;
    let DATA = EMBEDDED;
    let ACTIVE_DASHBOARD = null;

    function runGas(name, args = []) {
      return new Promise(resolve => {
        if (!window.google || !google.script || !google.script.run) return resolve(null);
        google.script.run.withSuccessHandler(resolve).withFailureHandler(() => resolve(null))[name].apply(google.script.run, args);
      });
    }
    function num(v) {
      const n = Number(String(v ?? '').replace(/[¥￥,%Q台件日\s]/g,''));
      return Number.isFinite(n) ? n : 0;
    }
    function pickValue(obj, keys, fallback = 0) {
      for (const key of keys) {
        if (obj && obj[key] !== undefined && obj[key] !== '') return obj[key];
      }
      return fallback;
    }
    function calcRate(numerator, denominator) {
      const n = num(numerator), d = num(denominator);
      if (!d) return 0;
      return Math.round((n / d) * 1000) / 10;
    }
    function calcForecast(actual, elapsedDays, monthDays) {
      const a = num(actual), e = num(elapsedDays), m = num(monthDays);
      if (!e || !m) return 0;
      return Math.round(a / e * m);
    }
    function pct(v) {
      const n = num(v);
      return Math.abs(n) <= 1 ? n * 100 : n;
    }
    function esc(v) {
      return String(v ?? '').replace(/[&<>"']/g, s => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#039;'}[s]));
    }
    function fmt(v, type = 'number', suffix = '') {
      const n = num(v);
      if (type === 'money') return `${n < 0 ? '-' : ''}¥${Math.abs(Math.round(n * 1000)).toLocaleString('ja-JP')}`;
      if (type === 'yen') return `${n < 0 ? '-' : ''}¥${Math.abs(Math.round(n)).toLocaleString('ja-JP')}`;
      if (type === 'percent') return `${pct(v).toFixed(1)}%`;
      if (type === 'days') return `${n.toFixed(1).replace(/\.0$/,'')}日`;
      return `${Math.round(n).toLocaleString('ja-JP')}${suffix}`;
    }
    function rowFor(list, dept) {
      return (list || []).find(r => r && r['営業部名'] === dept) || {};
    }
    function mergedRow(dept) {
      if (ACTIVE_DASHBOARD && ACTIVE_DASHBOARD.selectedDept === dept) {
        return {
          ...(ACTIVE_DASHBOARD.summary || {}),
          dash: {
            '対象月': ACTIVE_DASHBOARD.targetMonth,
            '更新日時': ACTIVE_DASHBOARD.updatedAt
          },
          sourceSheetName: ACTIVE_DASHBOARD.sourceSheetName
        };
      }
      return {
        ...rowFor(DATA.osRows, dept),
        ...rowFor(DATA.purchaseRows, dept),
        ...rowFor(DATA.salesRows, dept),
        dash: rowFor(DATA.dashRows, dept)
      };
    }
    function valueClass(v) { return num(v) < 0 ? 'negative' : ''; }
    function card(label, value, type='number', suffix='', cls='detail-card') {
      return `<article class="${cls}"><div class="label">${esc(label)}</div><div class="value ${valueClass(value)}">${fmt(value,type,suffix)}</div></article>`;
    }
    function ratioCard(label, q, rateValue, suffix='Q', color='') {
      const r = pct(rateValue);
      return `<article class="ratio-card">
        <div class="label">${esc(label)}</div>
        <div class="value">${fmt(q,'number',suffix)} / ${r.toFixed(1)}%</div>
        <div class="bar-meta"><span>${esc(label)}比率</span><span>${r.toFixed(1)}%</span></div>
        <div class="progress-bar"><div class="progress-fill ${color}" data-width="${Math.max(0,Math.min(100,r))}"></div></div>
      </article>`;
    }
    function progressCard(label, target, actual, elapsedDays, monthDays) {
      const forecast = calcForecast(actual, elapsedDays, monthDays);
      const rate = calcRate(forecast, target);
      const diff = forecast - num(target);
      const color = rate >= 100 ? 'green' : rate >= 80 ? 'warn' : 'bad';
      return `<article class="progress-card">
        <div class="label">${esc(label)}</div>
        <div class="value">${fmt(actual)} <span class="unit">現在</span></div>
        <div class="bar-meta"><span>目標 ${fmt(target)}</span><span>予想 ${fmt(forecast)} / ${rate.toFixed(1)}%</span></div>
        <div class="progress-bar"><div class="progress-fill ${color}" data-width="${Math.max(0,Math.min(140,rate))}"></div></div>
        <div class="bar-meta"><span>差分</span><span class="${diff < 0 ? 'negative' : ''}">${fmt(diff)}</span></div>
      </article>`;
    }
    function stack(title, items, total) {
      const colors = ['red','black','blue','yellow'];
      const totalNum = Math.max(1, num(total) || items.reduce((s,it)=>s+num(it.value),0));
      const segs = items.map((it,i)=>`<span class="seg ${colors[i%colors.length]}" data-width="${Math.max(0,Math.min(100,num(it.value)/totalNum*100))}"></span>`).join('');
      const legend = items.map((it,i)=>`<span><i class="dot ${colors[i%colors.length]}"></i>${esc(it.label)} ${fmt(it.value,'number','Q')} / ${(num(it.value)/totalNum*100).toFixed(1)}%</span>`).join('');
      return `<article class="detail-card"><div class="label">${esc(title)}</div><div class="stackbar">${segs}</div><div class="legend">${legend}</div></article>`;
    }
    function section(title, tag, body) {
      return `<section class="major-section"><div class="section-head"><div class="section-title">${esc(title)}</div><div class="section-tag">${esc(tag)}</div></div>${body}</section>`;
    }
    function renderOnePageDashboard() {
      const dept = document.getElementById('deptSelect').value;
      const r = mergedRow(dept);
      const dash = r.dash || {};
      document.getElementById('targetMonth').textContent = `対象月 ${pickValue(dash,['対象月'], ACTIVE_DASHBOARD?.targetMonth || '-')}`;
      document.getElementById('updatedAt').textContent = `更新 ${pickValue(dash,['更新日時'], ACTIVE_DASHBOARD?.updatedAt || '-')}`;
      const roleTotal = num(pickValue(r,['責任者'],pickValue(dash,['責任者人数'],0))) + num(pickValue(r,['買取営業'],pickValue(dash,['買取営業人数'],0))) + num(pickValue(r,['販売営業'],pickValue(dash,['販売営業人数'],0))) + num(pickValue(r,['事務'],pickValue(dash,['事務人数'],0)));
      const staffTotal = pickValue(dash,['スタッフ数'], pickValue(r,['スタッフ数'], roleTotal)) || roleTotal;
      const monthDays = pickValue(dash,['今月日数'],30);
      const elapsed = pickValue(dash,['経過日数'],1);
      const app = document.getElementById('app');
      app.innerHTML = `
        <section class="summary-hero">
          <div>
            <div class="dept-name">${esc(dept || '-')}</div>
            <div class="hero-note">基本情報・最重要KPI・買取・販売・集客情報を1ページに統合表示${ACTIVE_DASHBOARD?.sourceSheetName ? ` / 参照: ${esc(ACTIVE_DASHBOARD.sourceSheetName)}` : ''}</div>
          </div>
          <div class="staff-grid">
            ${card('スタッフ数', staffTotal, 'number','人','staff-card')}
            ${card('責任者', pickValue(r,['責任者'],pickValue(dash,['責任者人数'],0)), 'number','人','staff-card')}
            ${card('買取営業', pickValue(r,['買取営業'],pickValue(dash,['買取営業人数'],0)), 'number','人','staff-card')}
            ${card('販売営業', pickValue(r,['販売営業'],pickValue(dash,['販売営業人数'],0)), 'number','人','staff-card')}
            ${card('事務', pickValue(r,['事務'],pickValue(dash,['事務人数'],0)), 'number','人','staff-card')}
          </div>
        </section>

        ${section('最重要項目','G / MQ / F',`
          <div class="core-kpi-grid">
            ${card('G 最終利益', pickValue(dash,['G'],0), 'money','','kpi-card big')}
            ${card('MQ 粗利', pickValue(dash,['トータルMQ'],pickValue(r,['MQ'],0)), 'money','','kpi-card big')}
            ${card('F 経費', pickValue(dash,['トータルF'],pickValue(r,['F'],0)), 'money','','kpi-card big')}
          </div>
        `)}

        ${section('買取実績','査定 / 成約 / 販路',`
          <div class="section-grid-3">
            ${card('査定Q', pickValue(r,['査定Ｑ','査定Q','査定台数実績'],0), 'number','Q','kpi-card big')}
            ${card('買取Q', pickValue(r,['買取Ｑ','買取Q','買取台数実績'],0), 'number','Q','kpi-card big')}
            ${card('成約CVR', pickValue(r,['成約CVR','買取成約CVR'],0), 'percent','','kpi-card big')}
          </div>
          <div class="sub-section section-grid-4">
            ${ratioCard('AA', pickValue(r,['AAQ','AA台数'],0), pickValue(r,['AA比率'],0))}
            ${ratioCard('商品', pickValue(r,['商品Q','展示台数'],0), pickValue(r,['商品比率'],0))}
            ${ratioCard('スクラップ', pickValue(r,['スクラップQ','スクラップ台数'],0), pickValue(r,['スクラップ比率'],0))}
            ${ratioCard('代車', pickValue(r,['代車Q','代車台数'],0), pickValue(r,['代車比率'],0))}
          </div>
          <div class="sub-section section-grid-2">
            ${card('スクラップMQ', pickValue(r,['スクラップ粗利','スクラップMQ'],0), 'money','','mq-card')}
            ${card('KBMQ', pickValue(r,['KB納車MQ','KBMQ'],0), 'money','','mq-card')}
          </div>
        `)}

        ${section('AA詳細','処理状況 / 収益',`
          <div class="section-grid-4">
            ${card('落札Q', pickValue(r,['AA落札Q','落札台数'],0), 'number','Q','kpi-card big')}
            ${ratioCard('流札Q', pickValue(r,['AA流札Q','流札台数'],0), calcRate(pickValue(r,['AA流札Q'],0), pickValue(r,['AA落札Q'],0)), 'Q','warn')}
            ${ratioCard('取消Q', pickValue(r,['AA取消Q','取消台数'],0), calcRate(pickValue(r,['AA取消Q'],0), pickValue(r,['AA落札Q'],0)), 'Q','bad')}
            ${ratioCard('赤字Q', pickValue(r,['AA赤字Q'],0), calcRate(pickValue(r,['AA赤字Q'],0), pickValue(r,['AA落札Q'],0)), 'Q','bad')}
          </div>
          <div class="sub-section section-grid-4">
            ${card('未処理Q', pickValue(r,['AA未処理Q','未処理Q'],0), 'number','Q','detail-card')}
            ${card('持越Q', pickValue(r,['AA翌月Q','翌月処理予定Q','持越Q'],0), 'number','Q','detail-card')}
            ${card('落札MQ', pickValue(r,['当月AAMQ','AAMQ'],0), 'money','','mq-card')}
            ${card('AA@', pickValue(r,['AA@','AA＠','AA平均粗利'],0), 'money','','mq-card')}
            ${card('未処理MQ', pickValue(r,['AA未処理MQ'],0), 'money','','mq-card')}
            ${card('赤字金額', pickValue(r,['AA赤字金額'],0), 'money','','bad-card')}
          </div>
        `)}

        ${section('査定成約','契約 / タイミング',`
          <div class="section-grid-4">
            ${ratioCard('即決成約', pickValue(r,['即決契約数'],0), calcRate(pickValue(r,['即決契約数'],0), pickValue(r,['買取Ｑ'],0)), '件')}
            ${ratioCard('管理契約', pickValue(r,['管理契約数'],0), calcRate(pickValue(r,['管理契約数'],0), pickValue(r,['買取Ｑ'],0)), '件')}
            ${ratioCard('キャンセルQ', pickValue(r,['買取キャンセルQ','買取キャンセル数'],0), pickValue(r,['キャンセル比率'],0), 'Q','bad')}
            ${card('入庫平均日数', pickValue(r,['入庫平均日数'],0), 'days','','detail-card')}
          </div>
          <div class="sub-section section-grid-4">
            ${ratioCard('即日', pickValue(r,['即日査定'],0), pickValue(r,['即日査定比率'],calcRate(pickValue(r,['即日査定'],0),pickValue(r,['買取Ｑ'],0))))}
            ${ratioCard('過去当日', pickValue(r,['過去当日査定'],0), pickValue(r,['過去当日査定比率'],0))}
            ${ratioCard('過去後日', pickValue(r,['過去後日査定'],0), pickValue(r,['過去後日査定比率'],0))}
            ${ratioCard('後日', pickValue(r,['後日査定'],0), pickValue(r,['後日査定比率'],0))}
          </div>
        `)}

        ${section('販売 受注納車','納車 / 未納車 / 回転',`
          <div class="section-grid-4">
            ${card('納車台数', pickValue(r,['当月納車Ｑ','納車台数実績'],0), 'number','Q','kpi-card big')}
            ${card('未納車台数', pickValue(r,['当月未納車Ｑ'],0), 'number','Q','kpi-card big')}
            ${card('翌月納車Q', pickValue(r,['翌月納車Ｑ'],0), 'number','Q','detail-card')}
            ${ratioCard('即納Q', pickValue(r,['当月受注納車Q'],0), pickValue(r,['当月即納比率'],0))}
            ${ratioCard('前月Q', pickValue(r,['前月受注納車Q'],0), pickValue(r,['納車Q前月比率'],0))}
            ${ratioCard('当月比率', pickValue(r,['当月受注納車Q'],0), pickValue(r,['納車Q当月比率'],0))}
            ${ratioCard('翌月比率', pickValue(r,['翌月納車Ｑ'],0), pickValue(r,['納車Q翌月比率'],0))}
            ${card('回転率', calcRate(pickValue(r,['当月納車Ｑ'],0), pickValue(r,['在庫台数'],0)), 'percent','','detail-card')}
          </div>
        `)}

        ${section('販売 在庫状況','在庫 / 掲載 / キャパ',`
          <div class="section-grid-4">
            ${card('在庫Q', pickValue(r,['在庫台数'],0), 'number','Q','kpi-card big')}
            ${card('配車在庫Q', pickValue(r,['配車在庫数'],0), 'number','Q','kpi-card big')}
            ${ratioCard('キャパQ', pickValue(r,['キャパQ','在庫キャパ台数'],0), calcRate(pickValue(r,['在庫台数'],0), pickValue(r,['キャパQ','在庫キャパ台数'],0)), 'Q')}
            ${card('在庫金額', pickValue(r,['在庫金額'],0), 'money','','kpi-card big')}
          </div>
          <div class="sub-section section-grid-4">
            ${card('受注Q', pickValue(r,['当月受注Ｑ','受注Q'],0), 'number','Q','detail-card')}
            ${card('販売済Q', pickValue(r,['販売済Q','販売台数実績'],0), 'number','Q','detail-card')}
            ${card('掲載Q', pickValue(r,['掲載数','掲載Q'],0), 'number','Q','detail-card')}
            ${card('未掲載Q', pickValue(r,['未掲載数','未掲載Q'],0), 'number','Q','detail-card')}
            ${card('未入庫Q', pickValue(r,['未入庫台数','未入庫Q'],0), 'number','Q','detail-card')}
            ${card('未受入Q', pickValue(r,['未受入Q'],0), 'number','Q','detail-card')}
          </div>
        `)}

        ${section('販売収益','MQ / KB / 付帯',`
          <div class="section-grid-4">
            ${card('納車済MQ', pickValue(r,['当月納車MQ','納車済MQ'],0), 'money','','kpi-card big')}
            ${card('販売総MQ', num(pickValue(r,['当月納車MQ'],0)) + num(pickValue(r,['当月未納車MQ'],0)) + num(pickValue(r,['翌月納車MQ'],0)), 'money','','kpi-card big')}
            ${card('未納車MQ', pickValue(r,['当月未納車MQ'],0), 'money','','mq-card')}
            ${card('翌月納車MQ', pickValue(r,['翌月納車MQ'],0), 'money','','mq-card')}
            ${card('納車KBMQ', pickValue(r,['KB納車MQ'],0), 'money','','mq-card')}
            ${card('未納車KBMQ', pickValue(r,['KB未納車MQ'],0), 'money','','mq-card')}
            ${card('納車付帯MQ', pickValue(r,['当月付帯MQ'],0), 'money','','mq-card')}
            ${card('未納車付帯MQ', pickValue(r,['当月付帯未MQ'],0), 'money','','mq-card')}
          </div>
        `)}

        ${section('付帯詳細','Q / 比率 / MQ',`
          <div class="section-grid-4">
            ${accessoryCard(r,'クレジット')}
            ${accessoryCard(r,'保証')}
            ${accessoryCard(r,'メンテナンス')}
            ${accessoryCard(r,'クリーニング')}
            ${accessoryCard(r,'コーティング')}
            ${accessoryCard(r,'エアコン')}
            ${accessoryCard(r,'楽々納車')}
            ${accessoryCard(r,'タイヤ交換')}
          </div>
        `)}

        ${section('集客情報','情報数 / CVR / CPA',`
          <div class="section-grid-4">
            ${card('情報数', pickValue(r,['MOTA総数','買取情報数'],0), 'number','件','kpi-card big')}
            ${card('開示数', pickValue(r,['MOTA総開示数','開示数'],0), 'number','件','kpi-card big')}
            ${card('査定数', pickValue(r,['査定総数','査定Ｑ'],0), 'number','件','kpi-card big')}
            ${card('成約数', pickValue(r,['成約総数','総契約'],0), 'number','件','kpi-card big')}
            ${ratioCard('開示率', pickValue(r,['MOTA総開示数'],0), pickValue(r,['総開示率'],0),'件','blue')}
            ${ratioCard('査定率', pickValue(r,['査定総数'],0), pickValue(r,['査定率'],0),'件','blue')}
            ${ratioCard('成約率', pickValue(r,['成約総数'],0), pickValue(r,['総成約率'],0),'件','blue')}
          </div>
          <div class="sub-section section-grid-4">
            ${progressCard('情報数', pickValue(dash,['買取情報数目標'],0), pickValue(r,['MOTA総数','買取情報数'],0), elapsed, monthDays)}
            ${progressCard('開示数', pickValue(dash,['買取情報数目標'],0), pickValue(r,['MOTA総開示数'],0), elapsed, monthDays)}
            ${progressCard('査定数', pickValue(dash,['査定Q目標'],0), pickValue(r,['査定総数','査定Ｑ'],0), elapsed, monthDays)}
            ${progressCard('成約数', pickValue(dash,['買取Q目標'],0), pickValue(r,['成約総数','買取Ｑ'],0), elapsed, monthDays)}
          </div>
          <div class="sub-section section-grid-3">
            ${card('情報料金', pickValue(r,['総料金','情報料'],0), 'yen','','detail-card')}
            ${card('成約CPA', calcCpa(pickValue(r,['総料金','情報料'],0), pickValue(r,['成約総数','総契約'],0)), 'yen','','detail-card')}
            ${card('査定CPA', calcCpa(pickValue(r,['総料金','情報料'],0), pickValue(r,['査定総数','査定Ｑ'],0)), 'yen','','detail-card')}
          </div>
          <div class="sub-section section-grid-3">
            ${card('買取営業人数', pickValue(r,['買取営業'],pickValue(dash,['買取営業人数'],0)), 'number','人','detail-card')}
            ${card('一人当たり情報数', calcPerPerson(pickValue(r,['MOTA総数','買取情報数'],0), pickValue(r,['買取営業'],pickValue(dash,['買取営業人数'],0))), 'number','件','detail-card')}
            ${stack('温度別構成',[
              {label:'直近層',value:pickValue(r,['MOTA直近総数'],0)},
              {label:'検討層',value:pickValue(r,['MOTA検討総数'],0)},
              {label:'潜在層',value:pickValue(r,['MOTA顕在総数'],0)}
            ],pickValue(r,['MOTA総数'],0))}
          </div>
        `)}

        <p class="footer-note">不足候補: 未受入Q、キャパQ、納車Q翌月比率、各付帯MQ、AA未処理MQは現行OSシートにない場合0表示。DASH_DBヘッダー追加候補として管理。</p>
      `;
      runBars();
    }
    function accessoryCard(r,name) {
      return `<article class="ratio-card">
        <div class="label">${esc(name)}</div>
        <div class="value">${fmt(pickValue(r,[`${name}Q`,`${name}件数`],0),'number','Q')}</div>
        <div class="bar-meta"><span>${esc(name)}比率</span><span>${fmt(pickValue(r,[`${name}比率`],0),'percent')}</span></div>
        <div class="progress-bar"><div class="progress-fill blue" data-width="${Math.max(0,Math.min(100,pct(pickValue(r,[`${name}比率`],0))))}"></div></div>
        <div class="bar-meta"><span>MQ</span><span>${fmt(pickValue(r,[`${name}MQ`],0),'money')}</span></div>
      </article>`;
    }
    function calcCpa(cost, count) {
      const c = num(cost), n = num(count);
      return n ? Math.round(c / n) : 0;
    }
    function calcPerPerson(total, people) {
      const p = num(people);
      return p ? Math.round(num(total) / p) : 0;
    }
    function runBars() {
      document.querySelectorAll('.progress-fill,.seg').forEach(el => {
        el.style.width = '0%';
        const w = Number(el.dataset.width || 0);
        requestAnimationFrame(() => requestAnimationFrame(() => { el.style.width = `${Math.max(0,Math.min(100,w))}%`; }));
      });
    }
    async function loadDeptOptions() {
      const options = await runGas('getSalesOsDeptOptions');
      if (Array.isArray(options) && options.length) {
        const select = document.getElementById('deptSelect');
        const prev = select.value;
        select.innerHTML = options.map(n => `<option>${esc(n)}</option>`).join('');
        select.value = options.includes(prev) ? prev : (options.includes('千葉営業部') ? '千葉営業部' : options[0]);
        return true;
      }
      setupDeptSelect();
      return false;
    }
    async function loadSelectedDept() {
      const dept = document.getElementById('deptSelect').value;
      const app = document.getElementById('app');
      app.innerHTML = `<section class="major-section"><div class="section-title">読み込み中</div><p class="footer-note">${esc(dept)} のリンクOSデータを取得しています。</p></section>`;
      const data = await runGas('getSalesOsDashboardByDept', [dept]);
      if (data && data.summary) {
        ACTIVE_DASHBOARD = data;
      } else {
        ACTIVE_DASHBOARD = null;
      }
      renderOnePageDashboard();
    }
    async function reloadOnePage() {
      ACTIVE_DASHBOARD = null;
      const hasRemoteOptions = await loadDeptOptions();
      if (hasRemoteOptions) {
        await loadSelectedDept();
        return;
      }
      const live = await runGas('getSalesOsOnePageRows');
      if (live && live.osRows) DATA = live;
      setupDeptSelect();
      renderOnePageDashboard();
    }
    async function refreshDashboard() {
      await runGas('refreshSalesDash', []);
      await reloadOnePage();
    }
    function setupDeptSelect() {
      const select = document.getElementById('deptSelect');
      const prev = select.value;
      const names = [...new Set((DATA.osRows || []).map(r => r['営業部名']).filter(Boolean))];
      select.innerHTML = names.map(n => `<option>${esc(n)}</option>`).join('');
      select.value = names.includes(prev) ? prev : (names.includes('千葉営業部') ? '千葉営業部' : names[0] || '');
    }
    reloadOnePage();
  </script>
</body>
</html>
"""

OUT.write_text(html.replace("__DATA__", json.dumps(payload, ensure_ascii=False)), encoding="utf-8")
print(OUT)
