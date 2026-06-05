import json
import sys

import openpyxl


path = sys.argv[1]
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
print("sheets=", wb.sheetnames)
for ws in wb.worksheets:
    print("---", ws.title, ws.max_row, ws.max_column)
    for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
        print(json.dumps([str(x) if x is not None else "" for x in row[:24]], ensure_ascii=False))
